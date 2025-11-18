"""
Clean My Data Downloads Module

Comprehensive download and report generation for clean-my-data tool.
Generates two complete export formats:
1. Excel Report (.xlsx) - Multi-sheet detailed analysis
2. JSON Report (.json) - Complete hierarchical data structure

Agents covered:
- null-handler: Null value detection and handling
- outlier-remover: Outlier detection and removal
- type-fixer: Type conversion and fixing
- governance-checker: Governance compliance validation
- test-coverage-agent: Test coverage validation
"""

from typing import Dict, List, Any, Tuple
from datetime import datetime
import base64
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class CleanMyDataDownloads:
    """Handles comprehensive downloads for clean-my-data tool."""
    
    def __init__(self):
        """Initialize download generator."""
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.subheader_font = Font(bold=True, size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    def generate_downloads(
        self,
        agent_results: Dict[str, Any],
        analysis_id: str,
        execution_time_ms: int,
        alerts: List[Dict],
        issues: List[Dict],
        recommendations: List[Dict],
        cleaned_files: Dict[str, Dict[str, Any]] = None
    ) -> List[Dict[str, Any]]:
        """
        Generate both Excel and JSON downloads, plus cleaned data files from agents.
        
        Args:
            agent_results: Dictionary of agent_id -> agent output
            analysis_id: Unique analysis identifier
            execution_time_ms: Total execution time
            alerts: List of alerts
            issues: List of issues
            recommendations: List of recommendations
            cleaned_files: Dictionary of agent_id -> cleaned_file metadata (with base64 content)
        
        Returns:
            List of download metadata dicts with base64 encoded content.
        """
        downloads = []
        cleaned_files = cleaned_files or {}
        
        # Generate Excel report
        excel_data = self._generate_excel_report(
            agent_results=agent_results,
            analysis_id=analysis_id,
            execution_time_ms=execution_time_ms,
            alerts=alerts,
            issues=issues,
            recommendations=recommendations
        )
        downloads.append(excel_data)
        
        # Generate JSON report
        json_data = self._generate_json_report(
            agent_results=agent_results,
            analysis_id=analysis_id,
            execution_time_ms=execution_time_ms,
            alerts=alerts,
            issues=issues,
            recommendations=recommendations
        )
        downloads.append(json_data)
        
        # ==================== ADD CLEANED DATA FILES ====================
        # Add cleaned CSV files from individual agents
        agent_names = {
            "cleanse-previewer": "Cleanse Previewer",
            "null-handler": "Null Handler",
            "outlier-remover": "Outlier Remover",
            "type-fixer": "Type Fixer",
            "duplicate-resolver": "Duplicate Resolver",
            "field-standardization": "Field Standardization",
            "quarantine-agent": "Quarantine Agent"
        }
        
        for agent_id, agent_name in agent_names.items():
            if agent_id in cleaned_files:
                cleaned_file_data = cleaned_files[agent_id]
                
                # Create download entry for cleaned file following the standard pattern
                download_entry = {
                    "download_id": f"{analysis_id}_cleaned_{agent_id}",
                    "name": f"Clean My Data - {agent_name} Cleaned Data",
                    "format": "csv",
                    "file_name": cleaned_file_data.get("filename", f"cleaned_{agent_id}.csv"),
                    "description": f"Cleaned data file produced by {agent_name} agent with all cleaning operations applied",
                    "mimeType": "text/csv",
                    "content_base64": cleaned_file_data.get("content", ""),  # Already base64 encoded from agent
                    "size_bytes": cleaned_file_data.get("size_bytes", 0),
                    "creation_date": datetime.utcnow().isoformat() + "Z",
                    "agent_id": agent_id
                }
                
                downloads.append(download_entry)
        
        # ==================== ADD QUARANTINED DATA FILES ====================
        # Add quarantined records file from quarantine agent
        if "quarantine-agent" in cleaned_files:
            quarantine_file = cleaned_files.get("quarantine-agent", {}).get("quarantine_file", {})
            
            if quarantine_file:
                quarantine_entry = {
                    "download_id": f"{analysis_id}_quarantined_records",
                    "name": "Clean My Data - Quarantined Records",
                    "format": "csv",
                    "file_name": quarantine_file.get("filename", "quarantined_records.csv"),
                    "description": "Records identified and isolated by Quarantine Agent as invalid or suspicious. Contains problematic data with quarantine timestamps and reasons.",
                    "mimeType": "text/csv",
                    "content_base64": quarantine_file.get("content", ""),
                    "size_bytes": quarantine_file.get("size_bytes", 0),
                    "creation_date": datetime.utcnow().isoformat() + "Z",
                    "agent_id": "quarantine-agent",
                    "data_type": "quarantine_zone"
                }
                
                downloads.append(quarantine_entry)
        
        return downloads
    
    def _generate_excel_report(
        self,
        agent_results: Dict[str, Any],
        analysis_id: str,
        execution_time_ms: int,
        alerts: List[Dict],
        issues: List[Dict],
        recommendations: List[Dict]
    ) -> Dict[str, Any]:
        """Generate comprehensive Excel report with all agent data."""
        try:
            wb = Workbook()
            wb.remove(wb.active)
            
            # Extract agent outputs
            null_output = agent_results.get("null-handler", {})
            outlier_output = agent_results.get("outlier-remover", {})
            type_output = agent_results.get("type-fixer", {})
            duplicate_output = agent_results.get("duplicate-resolver", {})
            governance_output = agent_results.get("governance-checker", {})
            test_output = agent_results.get("test-coverage-agent", {})
            quarantine_output = agent_results.get("quarantine-agent", {})
            
            # 1. ANALYSIS SUMMARY SHEET
            self._create_analysis_summary_sheet(
                wb, analysis_id, execution_time_ms, alerts, issues, recommendations
            )
            
            # 2. CLEANSE PREVIEWER SHEET
            cleanse_previewer_output = agent_results.get("cleanse-previewer", {})
            if cleanse_previewer_output.get("status") == "success":
                self._create_cleanse_previewer_sheet(wb, cleanse_previewer_output)
            
            # 3. NULL HANDLER SHEET
            if null_output.get("status") == "success":
                self._create_null_handler_sheet(wb, null_output)
            
            # 4. OUTLIER REMOVER SHEET
            if outlier_output.get("status") == "success":
                self._create_outlier_sheet(wb, outlier_output)
            
            # 5. TYPE FIXER SHEET
            if type_output.get("status") == "success":
                self._create_type_fixer_sheet(wb, type_output)
            
            # 6. DUPLICATE RESOLVER SHEET
            if duplicate_output.get("status") == "success":
                self._create_duplicate_resolver_sheet(wb, duplicate_output)
            
            # 7. FIELD STANDARDIZATION SHEET
            field_standardization_output = agent_results.get("field-standardization", {})
            if field_standardization_output.get("status") == "success":
                self._create_field_standardization_sheet(wb, field_standardization_output)
            
            # 8. QUARANTINE AGENT SHEET
            if quarantine_output.get("status") == "success":
                self._create_quarantine_sheet(wb, quarantine_output)
            
            # 9. GOVERNANCE CHECKER SHEET
            if governance_output.get("status") == "success":
                self._create_governance_sheet(wb, governance_output)
            
            # 10. TEST COVERAGE SHEET
            if test_output.get("status") == "success":
                self._create_test_coverage_sheet(wb, test_output)
            
            # 11. ALERTS SHEET
            if alerts:
                self._create_alerts_sheet(wb, alerts)
            
            # 12. ISSUES SHEET
            if issues:
                self._create_issues_sheet(wb, issues)
            
            # 13. RECOMMENDATIONS SHEET
            if recommendations:
                self._create_recommendations_sheet(wb, recommendations)
            
            # Convert to bytes and base64
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            file_content = output.getvalue()
            
            return {
                "download_id": f"{analysis_id}_clean_excel",
                "name": "Clean My Data - Complete Analysis Report",
                "format": "xlsx",
                "file_name": "clean_my_data_analysis.xlsx",
                "description": "Comprehensive Excel report with all cleaning analysis data, agent results, and detailed metrics",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "content_base64": base64.b64encode(file_content).decode('utf-8'),
                "size_bytes": len(file_content),
                "creation_date": datetime.utcnow().isoformat() + "Z",
                "sheets": ["Summary", "Null Handler", "Outlier Remover", "Type Fixer", "Governance", "Test Coverage", "Alerts", "Issues", "Recommendations"]
            }
        except Exception as e:
            print(f"Error generating Excel report: {str(e)}")
            return {
                "download_id": f"{analysis_id}_clean_excel_error",
                "status": "error",
                "error": str(e)
            }
    
    def _create_analysis_summary_sheet(self, wb, analysis_id, execution_time_ms, alerts, issues, recommendations):
        """Create analysis summary sheet."""
        ws = wb.create_sheet("Summary", 0)
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 50
        
        row = 1
        # Title
        ws[f'A{row}'] = "CLEAN MY DATA - ANALYSIS SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 2
        
        # Metadata
        metadata = [
            ["Analysis ID", analysis_id],
            ["Tool", "Clean My Data"],
            ["Timestamp", datetime.utcnow().isoformat() + 'Z'],
            ["Execution Time (ms)", execution_time_ms],
            ["Report Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        # Statistics
        ws[f'A{row}'] = "ANALYSIS STATISTICS"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        stats = [
            ["Total Alerts", len(alerts)],
            ["Total Issues", len(issues)],
            ["Total Recommendations", len(recommendations)],
            ["High Severity Alerts", len([a for a in alerts if a.get('severity') == 'high'])],
            ["Critical Alerts", len([a for a in alerts if a.get('severity') == 'critical'])]
        ]
        
        for key, value in stats:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
    
    def _create_cleanse_previewer_sheet(self, wb, agent_output):
        """Create cleanse previewer detailed sheet."""
        ws = wb.create_sheet("Cleanse Preview", 1)
        self._set_column_widths(ws, [30, 20, 20, 20, 50])
        
        row = 1
        ws[f'A{row}'] = "CLEANSE PREVIEWER ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Rules Previewed", summary_metrics.get("total_rules_previewed", 0)],
            ["High Impact Rules", summary_metrics.get("high_impact_rules", 0)],
            ["Medium Impact Rules", summary_metrics.get("medium_impact_rules", 0)],
            ["Low Impact Rules", summary_metrics.get("low_impact_rules", 0)],
            ["Safe to Execute", "YES" if summary_metrics.get("safe_to_execute", True) else "NO"],
            ["Total Warnings", summary_metrics.get("total_warnings", 0)]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Preview scores
        ws[f'A{row}'] = "PREVIEW SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        preview_score = data.get("preview_score", {})
        metrics = preview_score.get("metrics", {})
        score_items = [
            ["Overall Score", preview_score.get("overall_score", 0)],
            ["Accuracy Score", metrics.get("accuracy_score", 0)],
            ["Safety Score", metrics.get("safety_score", 0)],
            ["Completeness Score", metrics.get("completeness_score", 0)],
            ["Successful Simulations", f"{metrics.get('successful_simulations', 0)}/{metrics.get('total_simulations', 0)}"]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Simulated results
        ws[f'A{row}'] = "PREVIEW RESULTS BY RULE"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        headers = ["Rule ID", "Description", "Impact Level", "Row Change %", "Risk Level"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        preview_analysis = data.get("preview_analysis", {})
        simulated_results = preview_analysis.get("simulated_results", [])
        
        for result in simulated_results:
            ws.cell(row=row, column=1, value=result.get("rule_id", ""))
            ws.cell(row=row, column=2, value=result.get("rule_description", ""))
            ws.cell(row=row, column=3, value=result.get("impact_level", "unknown").upper())
            
            changes = result.get("changes", {})
            ws.cell(row=row, column=4, value=f"{changes.get('row_change_percentage', 0):.2f}%")
            
            risk_assessment = result.get("risk_assessment", {})
            ws.cell(row=row, column=5, value=risk_assessment.get("risk_level", "unknown").upper())
            
            for col_idx in range(1, 6):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
        
        row += 1
        
        # Overall impact assessment
        ws[f'A{row}'] = "OVERALL IMPACT ASSESSMENT"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        overall_impact = preview_analysis.get("overall_impact", {})
        impact_items = [
            ["Execution Safety", preview_analysis.get("execution_safety", "UNKNOWN")],
            ["Statistical Confidence", f"{preview_analysis.get('statistical_confidence', 0):.1f}%"],
            ["Safe to Execute", "YES" if overall_impact.get("safe_to_execute", True) else "NO"]
        ]
        
        for key, value in impact_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Warnings
        warnings = overall_impact.get("warnings", [])
        if warnings:
            ws[f'A{row}'] = "WARNINGS"
            ws[f'A{row}'].font = Font(bold=True, size=10)
            ws[f'A{row}'].fill = self.subheader_fill
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            for warning in warnings:
                ws[f'A{row}'] = warning
                ws[f'A{row}'].border = self.border
                ws[f'A{row}'].alignment = self.left_alignment
                ws.merge_cells(f'A{row}:E{row}')
                row += 1
    
    def _create_null_handler_sheet(self, wb, agent_output):
        """Create null handler detailed sheet."""
        ws = wb.create_sheet("Null Handler", 1)
        self._set_column_widths(ws, [25, 15, 15, 15, 15, 15, 50])
        
        row = 1
        # Header
        ws[f'A{row}'] = "NULL HANDLER ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = self.header_fill
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        row += 2
        
        # Agent metadata
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Total Null Values Handled", summary_metrics.get("total_nulls_handled", 0)],
            ["Columns with Nulls", summary_metrics.get("columns_with_nulls", 0)],
            ["Records Processed", summary_metrics.get("records_processed", 0)]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Cleaning score details
        ws[f'A{row}'] = "CLEANING SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        cleaning_score = data.get("cleaning_score", {})
        score_items = [
            ["Overall Score", cleaning_score.get("overall_score", 0)],
            ["Quality Status", cleaning_score.get("quality_status", "unknown")],
            ["Completeness", cleaning_score.get("completeness", 0)],
            ["Consistency", cleaning_score.get("consistency", 0)],
            ["Accuracy", cleaning_score.get("accuracy", 0)]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Null summary by column
        ws[f'A{row}'] = "NULL SUMMARY BY COLUMN"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        headers = ["Column", "Null Count", "Null %", "Handling Method", "Imputation Value", "Status", "Notes"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        null_analysis = data.get("null_analysis", {})
        null_summary = null_analysis.get("null_summary", {})
        
        for col_name, col_data in null_summary.items():
            ws.cell(row=row, column=1, value=col_name)
            ws.cell(row=row, column=2, value=col_data.get("null_count", 0))
            ws.cell(row=row, column=3, value=f"{col_data.get('null_percentage', 0):.2f}%")
            ws.cell(row=row, column=4, value=col_data.get("handling_method", ""))
            ws.cell(row=row, column=5, value=col_data.get("imputation_value", ""))
            ws.cell(row=row, column=6, value=col_data.get("status", ""))
            ws.cell(row=row, column=7, value=col_data.get("notes", ""))
            
            for col_idx in range(1, 8):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
        
        row += 1
        
        # Recommendations
        ws[f'A{row}'] = "RECOMMENDATIONS"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        for rec in null_analysis.get("recommendations", []):
            ws[f'A{row}'] = rec.get("action", "")
            ws[f'B{row}'] = rec.get("reason", "")
            ws[f'C{row}'] = rec.get("priority", "medium")
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            ws[f'C{row}'].border = self.border
            ws[f'A{row}'].alignment = self.left_alignment
            row += 1
    
    def _create_outlier_sheet(self, wb, agent_output):
        """Create outlier remover detailed sheet."""
        ws = wb.create_sheet("Outlier Remover", 2)
        self._set_column_widths(ws, [25, 15, 15, 15, 15, 15, 50])
        
        row = 1
        ws[f'A{row}'] = "OUTLIER REMOVER ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        # Metadata
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Total Outliers Detected", summary_metrics.get("total_outliers_detected", 0)],
            ["Outliers Removed", summary_metrics.get("outliers_removed", 0)],
            ["Numeric Columns Analyzed", summary_metrics.get("numeric_columns_analyzed", 0)]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Outlier scores
        ws[f'A{row}'] = "OUTLIER DETECTION SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        outlier_score = data.get("outlier_score", {})
        score_items = [
            ["Overall Score", outlier_score.get("overall_score", 0)],
            ["Quality Status", outlier_score.get("quality_status", "unknown")],
            ["Detection Accuracy", outlier_score.get("detection_accuracy", 0)],
            ["Removal Safety", outlier_score.get("removal_safety", 0)]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Outlier summary by column
        ws[f'A{row}'] = "OUTLIER SUMMARY BY COLUMN"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        headers = ["Column", "Outlier Count", "Outlier %", "Method Used", "Threshold", "Action Taken", "Confidence"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        outlier_analysis = data.get("outlier_analysis", {})
        outlier_summary = outlier_analysis.get("outlier_summary", {})
        
        for col_name, col_data in outlier_summary.items():
            ws.cell(row=row, column=1, value=col_name)
            ws.cell(row=row, column=2, value=col_data.get("outlier_count", 0))
            ws.cell(row=row, column=3, value=f"{col_data.get('outlier_percentage', 0):.2f}%")
            ws.cell(row=row, column=4, value=col_data.get("method_used", ""))
            ws.cell(row=row, column=5, value=col_data.get("threshold", ""))
            ws.cell(row=row, column=6, value=col_data.get("action", ""))
            ws.cell(row=row, column=7, value=col_data.get("confidence_score", 0))
            
            for col_idx in range(1, 8):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _create_type_fixer_sheet(self, wb, agent_output):
        """Create type fixer detailed sheet."""
        ws = wb.create_sheet("Type Fixer", 3)
        self._set_column_widths(ws, [25, 20, 20, 20, 50])
        
        row = 1
        ws[f'A{row}'] = "TYPE FIXER ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Type Issues Fixed", summary_metrics.get("type_issues_fixed", 0)],
            ["Columns with Issues", summary_metrics.get("columns_with_type_issues", 0)],
            ["Conversion Success Rate", f"{summary_metrics.get('conversion_success_rate', 0):.2f}%"]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Type fixing scores
        ws[f'A{row}'] = "FIXING SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        fixing_score = data.get("fixing_score", {})
        score_items = [
            ["Overall Score", fixing_score.get("overall_score", 0)],
            ["Quality", fixing_score.get("quality", "unknown")],
            ["Accuracy", fixing_score.get("accuracy", 0)],
            ["Safety", fixing_score.get("safety", 0)]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Type summary by column
        ws[f'A{row}'] = "TYPE ANALYSIS BY COLUMN"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        headers = ["Column", "Current Type", "Suggested Type", "Conversion Status", "Issues/Notes"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        type_analysis = data.get("type_analysis", {})
        type_summary = type_analysis.get("type_summary", {})
        
        for col_name, col_data in type_summary.items():
            ws.cell(row=row, column=1, value=col_name)
            ws.cell(row=row, column=2, value=col_data.get("current_type", ""))
            ws.cell(row=row, column=3, value=col_data.get("suggested_type", ""))
            ws.cell(row=row, column=4, value=col_data.get("conversion_status", ""))
            ws.cell(row=row, column=5, value="; ".join(col_data.get("issues", [])))
            
            for col_idx in range(1, 6):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _create_quarantine_sheet(self, wb, agent_output):
        """Create quarantine agent detailed sheet."""
        ws = wb.create_sheet("Quarantine", 5)
        self._set_column_widths(ws, [25, 15, 15, 15, 15, 50])
        
        row = 1
        ws[f'A{row}'] = "QUARANTINE AGENT ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        # Metadata
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Total Rows Processed", summary_metrics.get("total_rows_processed", 0)],
            ["Quarantined Records", summary_metrics.get("quarantined_records", 0)],
            ["Clean Records", summary_metrics.get("clean_records", 0)],
            ["Quarantine %", f"{summary_metrics.get('quarantine_percentage', 0):.2f}%"]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Quality scores
        ws[f'A{row}'] = "QUARANTINE QUALITY SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        quality_score = data.get("quality_score", {})
        metrics = quality_score.get("metrics", {})
        score_items = [
            ["Overall Score", quality_score.get("overall_score", 0)],
            ["Quality Status", data.get("quality_status", "unknown")],
            ["Quarantine Reduction Rate", f"{metrics.get('quarantine_reduction_rate', 0):.1f}%"],
            ["Data Integrity Rate", f"{metrics.get('data_integrity_rate', 0):.1f}%"],
            ["Processing Efficiency", f"{metrics.get('processing_efficiency_rate', 0):.1f}%"]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Quarantine issues by type
        ws[f'A{row}'] = "ISSUES BY TYPE"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        headers = ["Issue Type", "Count", "Percentage"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        quarantine_analysis = data.get("quarantine_analysis", {})
        issue_types = quarantine_analysis.get("issue_types", {})
        total_issues = sum(issue_types.values())
        
        for issue_type, count in issue_types.items():
            ws.cell(row=row, column=1, value=issue_type.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{(count / total_issues * 100) if total_issues > 0 else 0:.1f}%")
            
            for col_idx in range(1, 4):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
        
        row += 1
        
        # Severity breakdown
        ws[f'A{row}'] = "SEVERITY BREAKDOWN"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        headers = ["Severity Level", "Count", "Percentage"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        severity_breakdown = quarantine_analysis.get("severity_breakdown", {})
        severity_total = sum(severity_breakdown.values())
        
        for severity, count in severity_breakdown.items():
            ws.cell(row=row, column=1, value=severity.upper())
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{(count / severity_total * 100) if severity_total > 0 else 0:.1f}%")
            
            for col_idx in range(1, 4):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _create_duplicate_resolver_sheet(self, wb, agent_output):
        """Create duplicate resolver detailed sheet."""
        ws = wb.create_sheet("Duplicates", 6)
        self._set_column_widths(ws, [30, 15, 50])
        
        row = 1
        ws[f'A{row}'] = "DUPLICATE RESOLVER ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Duplicates Detected", summary_metrics.get("duplicates_detected", 0)],
            ["Duplicates Resolved", summary_metrics.get("duplicates_resolved", 0)],
            ["Rows Removed", summary_metrics.get("rows_removed", 0)],
            ["Remaining Rows", summary_metrics.get("remaining_rows", 0)]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Deduplication scores
        ws[f'A{row}'] = "DEDUPLICATION SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        dedup_score = data.get("dedup_score", {})
        metrics = dedup_score.get("metrics", {})
        score_items = [
            ["Overall Score", dedup_score.get("overall_score", 0)],
            ["Dedup Reduction Rate", f"{metrics.get('dedup_reduction_rate', 0):.1f}%"],
            ["Data Retention Rate", f"{metrics.get('data_retention_rate', 0):.1f}%"],
            ["Duplicate % Before", f"{metrics.get('duplicate_percentage_before', 0):.2f}%"],
            ["Duplicate % After", f"{metrics.get('duplicate_percentage_after', 0):.2f}%"]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Duplicate detection methods summary
        ws[f'A{row}'] = "DETECTION METHODS"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        headers = ["Detection Method", "Count", "Percentage"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        duplicate_analysis = data.get("duplicate_analysis", {})
        for method, method_data in duplicate_analysis.get("duplicate_summary", {}).items():
            if isinstance(method_data, dict):
                ws.cell(row=row, column=1, value=method.replace('_', ' ').title())
                ws.cell(row=row, column=2, value=method_data.get("duplicate_count", 0))
                ws.cell(row=row, column=3, value=f"{method_data.get('duplicate_percentage', 0):.2f}%")
                
                for col_idx in range(1, 4):
                    ws.cell(row=row, column=col_idx).border = self.border
                    ws.cell(row=row, column=col_idx).alignment = self.left_alignment
                row += 1
    
    def _create_field_standardization_sheet(self, wb, agent_output):
        """Create field standardization detailed sheet."""
        ws = wb.create_sheet("Field Standardization", 7)
        self._set_column_widths(ws, [30, 15, 15, 15, 50])
        
        row = 1
        ws[f'A{row}'] = "FIELD STANDARDIZATION ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Columns Standardized", summary_metrics.get("columns_standardized", 0)],
            ["Values Changed", summary_metrics.get("values_changed", 0)],
            ["Variations Reduced", summary_metrics.get("variations_reduced", 0)],
            ["Total Rows", summary_metrics.get("total_rows_processed", 0)]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Standardization scores
        ws[f'A{row}'] = "STANDARDIZATION SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        standardization_score = data.get("standardization_score", {})
        metrics = standardization_score.get("metrics", {})
        score_items = [
            ["Overall Score", standardization_score.get("overall_score", 0)],
            ["Standardization Effectiveness", f"{metrics.get('standardization_effectiveness', 0):.1f}%"],
            ["Data Retention Rate", f"{metrics.get('data_retention_rate', 0):.1f}%"],
            ["Variations Before", metrics.get('variations_before', 0)],
            ["Variations After", metrics.get('variations_after', 0)],
            ["Variations Reduced", metrics.get('variations_reduced', 0)]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Column improvements
        ws[f'A{row}'] = "COLUMN IMPROVEMENTS"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        headers = ["Column", "Unique Values Reduced", "Case Variations Fixed", "Whitespace Fixed", "Improvement %"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        standardization_analysis = data.get("standardization_analysis", {})
        improvements = standardization_analysis.get("improvements", {})
        column_improvements = improvements.get("column_improvements", {})
        
        for col_name, col_data in column_improvements.items():
            ws.cell(row=row, column=1, value=col_name)
            ws.cell(row=row, column=2, value=col_data.get("unique_values_reduced", 0))
            ws.cell(row=row, column=3, value=col_data.get("case_variations_fixed", 0))
            ws.cell(row=row, column=4, value=col_data.get("whitespace_issues_fixed", 0))
            ws.cell(row=row, column=5, value=f"{col_data.get('improvement_percentage', 0):.2f}%")
            
            for col_idx in range(1, 6):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
        
        row += 1
        
        # Standardization operations applied
        ws[f'A{row}'] = "OPERATIONS APPLIED"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        operations = standardization_analysis.get("standardization_operations", {})
        for operation, applied in operations.items():
            ws[f'A{row}'] = operation.replace('_', ' ').title()
            ws[f'B{row}'] = "Yes" if applied else "No"
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
    
    def _create_governance_sheet(self, wb, agent_output):
        """Create governance checker detailed sheet."""
        ws = wb.create_sheet("Governance", 5)
        self._set_column_widths(ws, [30, 15, 50])
        
        row = 1
        ws[f'A{row}'] = "GOVERNANCE CHECKER ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Compliance Status", data.get("compliance_status", "unknown")],
            ["Governance Issues Found", summary_metrics.get("governance_issues_found", 0)],
            ["Fields Analyzed", summary_metrics.get("fields_analyzed", 0)]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Governance scores
        ws[f'A{row}'] = "GOVERNANCE SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        governance_scores = data.get("governance_scores", {})
        score_items = [
            ["Overall Score", governance_scores.get("overall", 0)],
            ["Lineage", governance_scores.get("lineage", 0)],
            ["Classification", governance_scores.get("classification", 0)],
            ["Consent Management", governance_scores.get("consent", 0)],
            ["Access Control", governance_scores.get("access_control", 0)]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Governance issues
        ws[f'A{row}'] = "GOVERNANCE ISSUES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        headers = ["Issue Type", "Severity", "Description"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        row += 1
        
        for issue in data.get("governance_issues", []):
            ws.cell(row=row, column=1, value=issue.get("issue_type", ""))
            ws.cell(row=row, column=2, value=issue.get("severity", ""))
            ws.cell(row=row, column=3, value=issue.get("description", ""))
            
            for col_idx in range(1, 4):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _create_test_coverage_sheet(self, wb, agent_output):
        """Create test coverage detailed sheet."""
        ws = wb.create_sheet("Test Coverage", 6)
        self._set_column_widths(ws, [30, 15, 50])
        
        row = 1
        ws[f'A{row}'] = "TEST COVERAGE ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Coverage Status", data.get("coverage_status", "unknown")],
            ["Test Issues Found", summary_metrics.get("test_issues_count", 0)],
            ["Fields Tested", summary_metrics.get("fields_tested", 0)]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Test scores
        ws[f'A{row}'] = "TEST COVERAGE SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        test_scores = data.get("test_coverage_scores", {})
        score_items = [
            ["Overall Score", test_scores.get("overall", 0)],
            ["Uniqueness Tests", test_scores.get("uniqueness", 0)],
            ["Range Tests", test_scores.get("range", 0)],
            ["Format Tests", test_scores.get("format", 0)],
            ["Consistency Tests", test_scores.get("consistency", 0)]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Test issues
        ws[f'A{row}'] = "TEST COVERAGE ISSUES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        headers = ["Test Type", "Severity", "Description"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        row += 1
        
        for issue in data.get("test_coverage_issues", []):
            ws.cell(row=row, column=1, value=issue.get("test_type", ""))
            ws.cell(row=row, column=2, value=issue.get("severity", ""))
            ws.cell(row=row, column=3, value=issue.get("description", ""))
            
            for col_idx in range(1, 4):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _create_alerts_sheet(self, wb, alerts):
        """Create alerts sheet."""
        ws = wb.create_sheet("Alerts", 6)
        self._set_column_widths(ws, [20, 15, 20, 50, 20, 30])
        
        row = 1
        headers = ["Alert ID", "Severity", "Category", "Message", "Affected Fields", "Recommendation"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        for alert in alerts:
            ws.cell(row=row, column=1, value=alert.get("alert_id", ""))
            ws.cell(row=row, column=2, value=alert.get("severity", ""))
            ws.cell(row=row, column=3, value=alert.get("category", ""))
            ws.cell(row=row, column=4, value=alert.get("message", ""))
            ws.cell(row=row, column=5, value=alert.get("affected_fields_count", ""))
            ws.cell(row=row, column=6, value=alert.get("recommendation", ""))
            
            for col_idx in range(1, 7):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _create_issues_sheet(self, wb, issues):
        """Create issues sheet."""
        ws = wb.create_sheet("Issues", 7)
        self._set_column_widths(ws, [20, 20, 20, 20, 15, 50])
        
        row = 1
        headers = ["Issue ID", "Agent", "Field", "Issue Type", "Severity", "Message"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        for issue in issues:
            ws.cell(row=row, column=1, value=issue.get("issue_id", ""))
            ws.cell(row=row, column=2, value=issue.get("agent_id", ""))
            ws.cell(row=row, column=3, value=issue.get("field_name", ""))
            ws.cell(row=row, column=4, value=issue.get("issue_type", ""))
            ws.cell(row=row, column=5, value=issue.get("severity", ""))
            ws.cell(row=row, column=6, value=issue.get("message", ""))
            
            for col_idx in range(1, 7):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _create_recommendations_sheet(self, wb, recommendations):
        """Create recommendations sheet."""
        ws = wb.create_sheet("Recommendations", 8)
        self._set_column_widths(ws, [25, 20, 20, 15, 50, 15])
        
        row = 1
        headers = ["Recommendation ID", "Agent", "Field", "Priority", "Recommendation", "Timeline"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        for rec in recommendations:
            ws.cell(row=row, column=1, value=rec.get("recommendation_id", ""))
            ws.cell(row=row, column=2, value=rec.get("agent_id", ""))
            ws.cell(row=row, column=3, value=rec.get("field_name", ""))
            ws.cell(row=row, column=4, value=rec.get("priority", ""))
            ws.cell(row=row, column=5, value=rec.get("recommendation", ""))
            ws.cell(row=row, column=6, value=rec.get("timeline", ""))
            
            for col_idx in range(1, 7):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _set_column_widths(self, ws, widths):
        """Set column widths for worksheet."""
        for col_idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    def _generate_json_report(
        self,
        agent_results: Dict[str, Any],
        analysis_id: str,
        execution_time_ms: int,
        alerts: List[Dict],
        issues: List[Dict],
        recommendations: List[Dict]
    ) -> Dict[str, Any]:
        """Generate comprehensive JSON report with all agent data."""
        try:
            report_data = {
                "metadata": {
                    "analysis_id": analysis_id,
                    "tool": "clean-my-data",
                    "timestamp": datetime.utcnow().isoformat() + "Z",
                    "execution_time_ms": execution_time_ms,
                    "report_version": "1.0"
                },
                "summary": {
                    "total_alerts": len(alerts),
                    "total_issues": len(issues),
                    "total_recommendations": len(recommendations),
                    "critical_alerts": len([a for a in alerts if a.get('severity') == 'critical']),
                    "high_severity_alerts": len([a for a in alerts if a.get('severity') == 'high']),
                    "medium_severity_alerts": len([a for a in alerts if a.get('severity') == 'medium'])
                },
                "alerts": alerts,
                "issues": issues,
                "recommendations": recommendations,
                "agent_results": {}
            }
            
            # Add complete agent outputs
            for agent_id, agent_output in agent_results.items():
                if agent_output.get("status") == "success":
                    report_data["agent_results"][agent_id] = {
                        "status": agent_output.get("status"),
                        "execution_time_ms": agent_output.get("execution_time_ms", 0),
                        "summary_metrics": agent_output.get("summary_metrics", {}),
                        "data": agent_output.get("data", {}),
                        "timestamp": agent_output.get("timestamp", "")
                    }
            
            # Convert to JSON string
            json_str = json.dumps(report_data, indent=2, default=str)
            json_bytes = json_str.encode('utf-8')
            
            return {
                "download_id": f"{analysis_id}_clean_json",
                "name": "Clean My Data - Complete Analysis JSON",
                "format": "json",
                "file_name": "clean_my_data_analysis.json",
                "description": "Complete hierarchical JSON report with all analysis data, including raw agent outputs and metrics",
                "mimeType": "application/json",
                "content_base64": base64.b64encode(json_bytes).decode('utf-8'),
                "size_bytes": len(json_bytes),
                "creation_date": datetime.utcnow().isoformat() + "Z"
            }
        except Exception as e:
            print(f"Error generating JSON report: {str(e)}")
            return {
                "download_id": f"{analysis_id}_clean_json_error",
                "status": "error",
                "error": str(e)
            }
