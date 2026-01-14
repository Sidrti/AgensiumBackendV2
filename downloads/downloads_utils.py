"""
Downloads Utilities Module

Shared utilities for Excel and JSON report generation across download modules.
Provides common styling, formatting, sheet creation, and the BaseDownloader class.
"""

from typing import Dict, List, Any
from datetime import datetime
import json
import base64
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelStyler:
    """Provides consistent Excel styling across all download modules."""
    
    def __init__(self):
        """Initialize Excel styles."""
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.subheader_font = Font(bold=True, size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    def set_column_widths(self, ws, widths: List[int]):
        """Set column widths for worksheet."""
        for col_idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width


class CommonSheetCreator:
    """Creates common sheets used across different download modules."""
    
    def __init__(self, styler: ExcelStyler):
        """Initialize with an ExcelStyler instance."""
        self.styler = styler
    
    def create_ai_summary_sheet(self, wb: Workbook, analysis_summary: Dict[str, Any]):
        """Create AI-generated analysis summary sheet."""
        ws = wb.create_sheet("AI Analysis Summary")
        self.styler.set_column_widths(ws, [35, 80])
        
        row = 1
        ws[f'A{row}'] = "AI-GENERATED ANALYSIS SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.styler.header_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 2
        
        # Metadata
        metadata = [
            ["Status", analysis_summary.get("status", "")],
            ["Model Used", analysis_summary.get("model_used", "")],
            ["Generation Time (ms)", analysis_summary.get("execution_time_ms", 0)]
        ]
        
        for key, value in metadata:
            ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 1
        
        # Summary text
        ws[f'A{row}'] = "SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.styler.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        summary_text = analysis_summary.get("summary", "")
        ws.cell(row=row, column=1, value=summary_text)
        ws.merge_cells(f'A{row}:B{row}')
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Set row height for wrapped text
        ws.row_dimensions[row].height = max(15 * (len(summary_text) // 100 + 1), 30)
    
    def create_row_level_issues_sheet(self, wb: Workbook, row_level_issues: List[Dict], issue_summary: Dict[str, Any]):
        """Create row-level issues sheet."""
        ws = wb.create_sheet("Row-Level Issues")
        self.styler.set_column_widths(ws, [15, 25, 20, 15, 15, 50, 30])
        
        row = 1
        ws[f'A{row}'] = "ROW-LEVEL ISSUES"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.styler.header_fill
        ws.merge_cells(f'A{row}:G{row}')
        row += 2
        
        # Issue summary statistics
        ws[f'A{row}'] = "ISSUE SUMMARY STATISTICS"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.styler.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        summary_stats = [
            ["Total Issues", issue_summary.get("total_issues", 0)],
            ["Affected Rows", issue_summary.get("affected_rows", 0)],
            ["Affected Columns", len(issue_summary.get("affected_columns", []))],
            ["Critical Severity", issue_summary.get("by_severity", {}).get("critical", 0)],
            ["Warning Severity", issue_summary.get("by_severity", {}).get("warning", 0)],
            ["Info Severity", issue_summary.get("by_severity", {}).get("info", 0)]
        ]
        
        for key, value in summary_stats:
            ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 1
        
        # Issues by type
        ws[f'A{row}'] = "ISSUES BY TYPE"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.styler.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        for issue_type, count in issue_summary.get("by_type", {}).items():
            ws.cell(row=row, column=1, value=issue_type)
            ws.cell(row=row, column=2, value=count)
            row += 1
        
        row += 2
        
        # Row-level issues table
        ws[f'A{row}'] = "DETAILED ROW-LEVEL ISSUES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.styler.subheader_fill
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        headers = ["Row Index", "Column", "Issue Type", "Severity", "Agent ID", "Description", "Suggested Action"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.styler.header_fill
            cell.font = self.styler.header_font
            cell.border = self.styler.border
            cell.alignment = self.styler.center_alignment
        row += 1
        
        # Add issues (limit to first 1000)
        for issue in row_level_issues[:1000]:
            ws.cell(row=row, column=1, value=issue.get("row_index", ""))
            ws.cell(row=row, column=2, value=issue.get("column", ""))
            ws.cell(row=row, column=3, value=issue.get("issue_type", ""))
            ws.cell(row=row, column=4, value=issue.get("severity", ""))
            ws.cell(row=row, column=5, value=issue.get("agent_id", ""))
            ws.cell(row=row, column=6, value=issue.get("description", ""))
            ws.cell(row=row, column=7, value=issue.get("suggested_action", ""))
            
            for col_idx in range(1, 8):
                ws.cell(row=row, column=col_idx).border = self.styler.border
                ws.cell(row=row, column=col_idx).alignment = self.styler.left_alignment
            row += 1
    
    def create_routing_decisions_sheet(self, wb: Workbook, routing_decisions: List[Dict]):
        """Create routing decisions sheet."""
        ws = wb.create_sheet("Routing Decisions")
        self.styler.set_column_widths(ws, [20, 15, 20, 50, 30, 15])
        
        row = 1
        ws[f'A{row}'] = "ROUTING DECISIONS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.styler.header_fill
        ws.merge_cells(f'A{row}:F{row}')
        row += 2
        
        ws[f'A{row}'] = f"Total Routing Recommendations: {len(routing_decisions)}"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws.merge_cells(f'A{row}:B{row}')
        row += 2
        
        headers = ["Tool ID", "Priority", "Trigger Reason", "Rationale", "Expected Benefit", "Confidence"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.styler.header_fill
            cell.font = self.styler.header_font
            cell.border = self.styler.border
            cell.alignment = self.styler.center_alignment
        row += 1
        
        for decision in routing_decisions:
            ws.cell(row=row, column=1, value=decision.get("tool_id", ""))
            ws.cell(row=row, column=2, value=decision.get("priority", ""))
            ws.cell(row=row, column=3, value=decision.get("trigger_reason", ""))
            ws.cell(row=row, column=4, value=decision.get("rationale", ""))
            ws.cell(row=row, column=5, value=decision.get("expected_benefit", ""))
            ws.cell(row=row, column=6, value=decision.get("confidence", ""))
            
            for col_idx in range(1, 7):
                ws.cell(row=row, column=col_idx).border = self.styler.border
                ws.cell(row=row, column=col_idx).alignment = self.styler.left_alignment
            row += 1
    
    def create_alerts_sheet(self, wb: Workbook, alerts: List[Dict]):
        """Create alerts sheet."""
        ws = wb.create_sheet("Alerts")
        self.styler.set_column_widths(ws, [20, 15, 20, 50, 20, 30])
        
        row = 1
        headers = ["Alert ID", "Severity", "Category", "Message", "Affected Fields", "Recommendation"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.styler.header_fill
            cell.font = self.styler.header_font
            cell.border = self.styler.border
            cell.alignment = self.styler.center_alignment
        row += 1
        
        for alert in alerts:
            ws.cell(row=row, column=1, value=alert.get("alert_id", ""))
            ws.cell(row=row, column=2, value=alert.get("severity", ""))
            ws.cell(row=row, column=3, value=alert.get("category", ""))
            ws.cell(row=row, column=4, value=alert.get("message", ""))
            ws.cell(row=row, column=5, value=", ".join(alert.get("affected_fields", [])))
            ws.cell(row=row, column=6, value=alert.get("recommendation", ""))
            
            for col_idx in range(1, 7):
                ws.cell(row=row, column=col_idx).border = self.styler.border
                ws.cell(row=row, column=col_idx).alignment = self.styler.left_alignment
            row += 1
    
    def create_issues_sheet(self, wb: Workbook, issues: List[Dict]):
        """Create issues sheet."""
        ws = wb.create_sheet("Issues")
        self.styler.set_column_widths(ws, [20, 20, 20, 20, 15, 50])
        
        row = 1
        headers = ["Issue ID", "Agent", "Field", "Issue Type", "Severity", "Message"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.styler.header_fill
            cell.font = self.styler.header_font
            cell.border = self.styler.border
            cell.alignment = self.styler.center_alignment
        row += 1
        
        for issue in issues:
            ws.cell(row=row, column=1, value=issue.get("issue_id", ""))
            ws.cell(row=row, column=2, value=issue.get("agent_id", ""))
            ws.cell(row=row, column=3, value=issue.get("field_name", ""))
            ws.cell(row=row, column=4, value=issue.get("issue_type", ""))
            ws.cell(row=row, column=5, value=issue.get("severity", ""))
            ws.cell(row=row, column=6, value=issue.get("message", ""))
            
            for col_idx in range(1, 7):
                ws.cell(row=row, column=col_idx).border = self.styler.border
                ws.cell(row=row, column=col_idx).alignment = self.styler.left_alignment
            row += 1
    
    def create_recommendations_sheet(self, wb: Workbook, recommendations: List[Dict]):
        """Create recommendations sheet."""
        ws = wb.create_sheet("Recommendations")
        self.styler.set_column_widths(ws, [25, 20, 20, 15, 50, 15])
        
        row = 1
        headers = ["Recommendation ID", "Agent", "Field", "Priority", "Recommendation", "Timeline"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.styler.header_fill
            cell.font = self.styler.header_font
            cell.border = self.styler.border
            cell.alignment = self.styler.center_alignment
        row += 1
        
        for rec in recommendations:
            ws.cell(row=row, column=1, value=rec.get("recommendation_id", ""))
            ws.cell(row=row, column=2, value=rec.get("agent_id", ""))
            ws.cell(row=row, column=3, value=rec.get("field_name", ""))
            ws.cell(row=row, column=4, value=rec.get("priority", ""))
            ws.cell(row=row, column=5, value=rec.get("recommendation", ""))
            ws.cell(row=row, column=6, value=rec.get("timeline", ""))
            
            for col_idx in range(1, 7):
                ws.cell(row=row, column=col_idx).border = self.styler.border
                ws.cell(row=row, column=col_idx).alignment = self.styler.left_alignment
            row += 1


class BaseDownloader:
    """Base class for all tool-specific download handlers."""
    
    def __init__(self, tool_id: str, tool_display_name: str):
        self.tool_id = tool_id
        self.tool_display_name = tool_display_name
        self.styler = ExcelStyler()
        self.common_sheets = CommonSheetCreator(self.styler)
        
        # Legacy compat
        self.header_fill = self.styler.header_fill
        self.header_font = self.styler.header_font
        self.subheader_fill = self.styler.subheader_fill
        self.subheader_font = self.styler.subheader_font
        self.border = self.styler.border
        self.center_alignment = self.styler.center_alignment
        self.left_alignment = self.styler.left_alignment

    def generate_downloads(
        self,
        agent_results: Dict[str, Any],
        analysis_id: str,
        execution_time_ms: int,
        alerts: List[Dict],
        issues: List[Dict],
        recommendations: List[Dict],
        cleaned_files: Dict[str, Dict[str, Any]] = None,
        executive_summary: List[Dict] = None,
        analysis_summary: Dict[str, Any] = None,
        row_level_issues: List[Dict] = None,
        issue_summary: Dict[str, Any] = None,
        routing_decisions: List[Dict] = None
    ) -> List[Dict[str, Any]]:
        """
        Generate both Excel and JSON downloads.
        """
        cleaned_files = cleaned_files or {}
        executive_summary = executive_summary or []
        analysis_summary = analysis_summary or {}
        row_level_issues = row_level_issues or []
        issue_summary = issue_summary or {}
        routing_decisions = routing_decisions or []
        downloads = []
        
        # Generate Excel report
        excel_data = self._generate_excel_report(
            agent_results=agent_results,
            analysis_id=analysis_id,
            execution_time_ms=execution_time_ms,
            alerts=alerts,
            issues=issues,
            recommendations=recommendations,
            executive_summary=executive_summary,
            analysis_summary=analysis_summary,
            row_level_issues=row_level_issues,
            issue_summary=issue_summary,
            routing_decisions=routing_decisions
        )
        downloads.append(excel_data)
        
        # Generate JSON report
        json_data = self._generate_json_report(
            agent_results=agent_results,
            analysis_id=analysis_id,
            execution_time_ms=execution_time_ms,
            alerts=alerts,
            issues=issues,
            recommendations=recommendations,
            executive_summary=executive_summary,
            analysis_summary=analysis_summary,
            row_level_issues=row_level_issues,
            issue_summary=issue_summary,
            routing_decisions=routing_decisions
        )
        downloads.append(json_data)
        
        # Add cleaned/mastered files if present
        self._attach_cleaned_files(downloads, analysis_id, cleaned_files)
        
        return downloads

    def _generate_excel_report(self, **kwargs) -> Dict[str, Any]:
        """Orchestrate Excel creation. Subclasses implement create_tool_specific_sheets."""
        try:
            wb = Workbook()
            wb.remove(wb.active)
            
            # 1. Summary Sheet
            self._create_analysis_summary_sheet(wb, kwargs)
            
            # 2. Tool Specific Sheets (Abstract)
            self.create_tool_specific_sheets(wb, kwargs['agent_results'])
            
            # 3. Standard Sheets
            if kwargs.get('alerts'):
                self.common_sheets.create_alerts_sheet(wb, kwargs['alerts'])
            if kwargs.get('issues'):
                self.common_sheets.create_issues_sheet(wb, kwargs['issues'])
            if kwargs.get('recommendations'):
                self.common_sheets.create_recommendations_sheet(wb, kwargs['recommendations'])
            if kwargs.get('analysis_summary'):
                self.common_sheets.create_ai_summary_sheet(wb, kwargs['analysis_summary'])
            if kwargs.get('row_level_issues'):
                self.common_sheets.create_row_level_issues_sheet(wb, kwargs['row_level_issues'], kwargs.get('issue_summary', {}))
            if kwargs.get('routing_decisions'):
                self.common_sheets.create_routing_decisions_sheet(wb, kwargs['routing_decisions'])
            
            # Save
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            file_content = output.getvalue()
            
            tool_slug = self.tool_id.split('-')[0] # e.g. "clean" from "clean-my-data"
            return {
                "download_id": f"{kwargs['analysis_id']}_{tool_slug}_excel",
                "name": f"{self.tool_display_name} - Complete Analysis Report",
                "format": "xlsx",
                "file_name": f"{self.tool_id.replace('-', '_')}_analysis.xlsx",
                "description": f"Comprehensive Excel report for {self.tool_display_name}",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "content_base64": base64.b64encode(file_content).decode('utf-8'),
                "size_bytes": len(file_content),
                "creation_date": datetime.utcnow().isoformat() + "Z",
                "type": "complete_report",
                "sheets": wb.sheetnames
            }
        except Exception as e:
            print(f"Error generating Excel report: {str(e)}")
            return {
                "download_id": f"{kwargs['analysis_id']}_excel_error",
                "status": "error",
                "error": str(e)
            }

    def _create_analysis_summary_sheet(self, wb, kwargs):
        """Create standard analysis summary sheet."""
        ws = wb.create_sheet("Summary", 0)
        self.styler.set_column_widths(ws, [35, 50])
        
        row = 1
        ws[f'A{row}'] = f"{self.tool_display_name.upper()} - ANALYSIS SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f'A{row}'].fill = self.styler.header_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 2
        
        agents_names, _ = get_agents_metadata(self.tool_id, kwargs['agent_results'])
        
        metadata = [
            ["Analysis ID", kwargs['analysis_id']],
            ["Tool", self.tool_display_name],
            ["Timestamp", datetime.utcnow().isoformat() + 'Z'],
            ["Execution Time (ms)", kwargs['execution_time_ms']],
            ["Report Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")],
            ["Agents Used", ", ".join(agents_names)],
            ["Lineage", " -> ".join(agents_names)]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.styler.subheader_fill
            ws[f'A{row}'].border = self.styler.border
            ws[f'B{row}'].border = self.styler.border
            row += 1
        
        row += 1
        
        # Executive Summary
        if kwargs.get('executive_summary'):
            ws[f'A{row}'] = "EXECUTIVE SUMMARY"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            ws[f'A{row}'].fill = self.styler.subheader_fill
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            for item in kwargs['executive_summary']:
                ws[f'A{row}'] = item.get('title', '')
                ws[f'B{row}'] = f"{item.get('value', '')} - {item.get('description', '')}"
                ws[f'A{row}'].border = self.styler.border
                ws[f'B{row}'].border = self.styler.border
                row += 1
            row += 1
            
        # Statistics
        ws[f'A{row}'] = "ANALYSIS STATISTICS"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].fill = self.styler.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        stats = [
            ["Total Alerts", len(kwargs.get('alerts', []))],
            ["Total Issues", len(kwargs.get('issues', []))],
            ["Total Recommendations", len(kwargs.get('recommendations', []))]
        ]
        for key, value in stats:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.styler.border
            ws[f'B{row}'].border = self.styler.border
            row += 1

    def _generate_json_report(self, **kwargs) -> Dict[str, Any]:
        """Standard JSON report generation."""
        report_data = build_json_report_structure(
            analysis_id=kwargs['analysis_id'],
            tool=self.tool_id,
            execution_time_ms=kwargs['execution_time_ms'],
            alerts=kwargs['alerts'],
            issues=kwargs['issues'],
            recommendations=kwargs['recommendations'],
            executive_summary=kwargs['executive_summary'],
            analysis_summary=kwargs['analysis_summary'],
            row_level_issues=kwargs['row_level_issues'],
            issue_summary=kwargs['issue_summary'],
            routing_decisions=kwargs['routing_decisions'],
            agent_results=kwargs['agent_results']
        )
        
        tool_slug = self.tool_id.split('-')[0]
        return generate_json_download(
            analysis_id=kwargs['analysis_id'],
            tool=tool_slug,
            file_name=f"{self.tool_id.replace('-', '_')}_analysis.json",
            description=f"Complete hierarchical JSON report for {self.tool_display_name}",
            report_data=report_data
        )

    def _attach_cleaned_files(self, downloads: List[Dict], analysis_id: str, cleaned_files: Dict[str, Any]):
        """Attach cleaned/mastered files to the download list."""
        if not cleaned_files:
            return

        # Simple logic: assume the caller has already selected the "best" file and put it in cleaned_files
        # The key in cleaned_files is the agent_id.
        
        for agent_id, file_data in cleaned_files.items():
            if not file_data or not file_data.get("content"):
                continue
                
            download_entry = {
                "download_id": f"{analysis_id}_final_data_{agent_id}",
                "name": f"{self.tool_display_name} - Final Output",
                "format": file_data.get("format", "csv"),
                "file_name": file_data.get("filename", "final_data.csv"),
                "description": f"Final processed data file from {self.tool_display_name}",
                "mimeType": "text/csv", # Assuming CSV for now
                "content_base64": file_data.get("content", ""),
                "size_bytes": file_data.get("size_bytes", 0),
                "creation_date": datetime.utcnow().isoformat() + "Z",
                "type": "cleaned_data",
                "agent_id": agent_id
            }
            downloads.append(download_entry)

    def create_tool_specific_sheets(self, wb: Workbook, agent_results: Dict[str, Any]):
        """Abstract method to be implemented by subclasses."""
        pass


def load_tool_config(tool_id: str) -> Dict[str, Any]:
    """Load tool configuration from JSON file."""
    try:
        current_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(current_dir)
        tools_dir = os.path.join(project_root, 'tools')
        tool_file = os.path.join(tools_dir, f"{tool_id.replace('-', '_')}_tool.json")
        with open(tool_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading tool config for {tool_id}: {str(e)}")
        return {}


def get_agents_metadata(tool_id: str, agent_results: Dict[str, Any]) -> tuple:
    """Extract agent names and lineage."""
    tool_config = load_tool_config(tool_id)
    
    if not tool_config:
        agents_ids = [aid for aid, result in agent_results.items() if result.get("status") == "success"]
        agents_names = [result.get("agent_name", aid) for aid, result in agent_results.items() if result.get("status") == "success"]
        return agents_names, agents_ids
    
    available_agents = tool_config.get('tool', {}).get('available_agents', [])
    agents_config = tool_config.get('agents', {})
    
    agents_ids = [aid for aid in available_agents if aid in agent_results and agent_results[aid].get("status") == "success"]
    
    agents_names = []
    for aid in agents_ids:
        agent_name = agents_config.get(aid, {}).get('name')
        if not agent_name:
            agent_name = agent_results[aid].get('agent_name', aid)
        agents_names.append(agent_name)
    
    return agents_names, agents_ids


def build_json_report_structure(
    analysis_id: str,
    tool: str,
    execution_time_ms: int,
    alerts: List[Dict],
    issues: List[Dict],
    recommendations: List[Dict],
    executive_summary: List[Dict],
    analysis_summary: Dict[str, Any],
    row_level_issues: List[Dict],
    issue_summary: Dict[str, Any],
    routing_decisions: List[Dict],
    agent_results: Dict[str, Any]
) -> Dict[str, Any]:
    """Build standardized JSON report structure."""
    agents_names, agents_ids = get_agents_metadata(tool, agent_results)
    
    report_data = {
        "metadata": {
            "analysis_id": analysis_id,
            "tool": tool,
            "timestamp": datetime.utcnow().isoformat() + "Z",
            "execution_time_ms": execution_time_ms,
            "report_version": "1.0",
            "agents_name": agents_names,
            "lineage": agents_names
        },
        "executive_summary": executive_summary,
        "analysis_summary": analysis_summary,
        "summary": {
            "total_alerts": len(alerts),
            "total_issues": len(issues),
            "total_recommendations": len(recommendations),
            "critical_alerts": len([a for a in alerts if a.get('severity') == 'critical']),
            "high_severity_alerts": len([a for a in alerts if a.get('severity') == 'high']),
            "medium_severity_alerts": len([a for a in alerts if a.get('severity') == 'medium'])
        },
        "alerts": alerts,
        "issues": issues,
        "recommendations": recommendations,
        "row_level_issues": row_level_issues,
        "issue_summary": issue_summary,
        "routing_decisions": routing_decisions,
        "agent_results": {}
    }
    
    for agent_id, agent_output in agent_results.items():
        if agent_output.get("status") == "success":
            # Strip potentially large file content from JSON report
            sanitized_data = agent_output.get("data", {})
            if isinstance(agent_output, dict):
                sanitized_output = {k: v for k, v in agent_output.items() if k != 'cleaned_file'}
            else:
                sanitized_output = agent_output

            report_data["agent_results"][agent_id] = {
                "status": sanitized_output.get("status"),
                "execution_time_ms": sanitized_output.get("execution_time_ms", 0),
                "summary_metrics": sanitized_output.get("summary_metrics", {}),
                "data": sanitized_data,
                "timestamp": sanitized_output.get("timestamp", "")
            }
    
    return report_data


def generate_json_download(
    analysis_id: str,
    tool: str,
    file_name: str,
    description: str,
    report_data: Dict[str, Any]
) -> Dict[str, Any]:
    """Generate JSON download metadata."""
    try:
        json_str = json.dumps(report_data, indent=2, default=str)
        json_bytes = json_str.encode('utf-8')
        
        return {
            "download_id": f"{analysis_id}_{tool}_json",
            "name": f"{tool.replace('-', ' ').title()} - Complete Analysis JSON",
            "format": "json",
            "file_name": file_name,
            "description": description,
            "mimeType": "application/json",
            "content_base64": base64.b64encode(json_bytes).decode('utf-8'),
            "size_bytes": len(json_bytes),
            "creation_date": datetime.utcnow().isoformat() + "Z",
            "type": "complete_report"
        }
    except Exception as e:
        print(f"Error generating JSON download: {str(e)}")
        return {
            "download_id": f"{analysis_id}_{tool}_json_error",
            "status": "error",
            "error": str(e)
        }
