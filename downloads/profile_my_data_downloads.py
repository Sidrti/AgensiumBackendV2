"""
Profile My Data Downloads Module

Comprehensive download and report generation for profile-my-data tool.
Generates two complete export formats:
1. Excel Report (.xlsx) - Multi-sheet detailed analysis
2. JSON Report (.json) - Complete hierarchical data structure

Agents covered:
- unified-profiler: Field-level quality metrics
- drift-detector: Distribution changes from baseline
- score-risk: PII and compliance risk assessment
- readiness-rater: Overall data readiness scores
- governance-checker: Governance compliance validation
- test-coverage-agent: Test coverage validation
"""

from typing import Dict, List, Any
from datetime import datetime
import base64
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ProfileMyDataDownloads:
    """Handles comprehensive downloads for profile-my-data tool."""
    
    def __init__(self):
        """Initialize download generator."""
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.subheader_font = Font(bold=True, size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    def generate_downloads(
        self,
        agent_results: Dict[str, Any],
        analysis_id: str,
        execution_time_ms: int,
        alerts: List[Dict],
        issues: List[Dict],
        recommendations: List[Dict],
        executive_summary: List[Dict]
    ) -> List[Dict[str, Any]]:
        """
        Generate both Excel and JSON downloads.
        
        Returns list of download metadata dicts with base64 encoded content.
        """
        downloads = []
        
        # Generate Excel report
        excel_data = self._generate_excel_report(
            agent_results=agent_results,
            analysis_id=analysis_id,
            execution_time_ms=execution_time_ms,
            alerts=alerts,
            issues=issues,
            recommendations=recommendations,
            executive_summary=executive_summary
        )
        downloads.append(excel_data)
        
        # Generate JSON report
        json_data = self._generate_json_report(
            agent_results=agent_results,
            analysis_id=analysis_id,
            execution_time_ms=execution_time_ms,
            alerts=alerts,
            issues=issues,
            recommendations=recommendations,
            executive_summary=executive_summary
        )
        downloads.append(json_data)
        
        return downloads
    
    def _generate_excel_report(
        self,
        agent_results: Dict[str, Any],
        analysis_id: str,
        execution_time_ms: int,
        alerts: List[Dict],
        issues: List[Dict],
        recommendations: List[Dict],
        executive_summary: List[Dict]
    ) -> Dict[str, Any]:
        """Generate comprehensive Excel report with all agent data."""
        try:
            wb = Workbook()
            wb.remove(wb.active)
            
            # Extract agent outputs
            profiler_output = agent_results.get("unified-profiler", {})
            drift_output = agent_results.get("drift-detector", {})
            risk_output = agent_results.get("score-risk", {})
            readiness_output = agent_results.get("readiness-rater", {})
            governance_output = agent_results.get("governance-checker", {})
            test_output = agent_results.get("test-coverage-agent", {})
            
            # 1. ANALYSIS SUMMARY SHEET
            self._create_analysis_summary_sheet(
                wb, analysis_id, execution_time_ms, alerts, issues, recommendations, executive_summary
            )
            
            # 2. UNIFIED PROFILER SHEET
            if profiler_output.get("status") == "success":
                self._create_profiler_sheet(wb, profiler_output)
            
            # 3. DRIFT DETECTOR SHEET
            if drift_output.get("status") == "success":
                self._create_drift_sheet(wb, drift_output)
            
            # 4. SCORE RISK SHEET
            if risk_output.get("status") == "success":
                self._create_risk_sheet(wb, risk_output)
            
            # 5. READINESS RATER SHEET
            if readiness_output.get("status") == "success":
                self._create_readiness_sheet(wb, readiness_output)
            
            # 6. GOVERNANCE CHECKER SHEET
            if governance_output.get("status") == "success":
                self._create_governance_sheet(wb, governance_output)
            
            # 7. TEST COVERAGE SHEET
            if test_output.get("status") == "success":
                self._create_test_coverage_sheet(wb, test_output)
            
            # 8. ALERTS SHEET
            if alerts:
                self._create_alerts_sheet(wb, alerts)
            
            # 9. ISSUES SHEET
            if issues:
                self._create_issues_sheet(wb, issues)
            
            # 10. RECOMMENDATIONS SHEET
            if recommendations:
                self._create_recommendations_sheet(wb, recommendations)
            
            # Convert to bytes and base64
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            file_content = output.getvalue()
            
            return {
                "download_id": f"{analysis_id}_profile_excel",
                "name": "Profile My Data - Complete Analysis Report",
                "format": "xlsx",
                "file_name": "profile_my_data_analysis.xlsx",
                "description": "Comprehensive Excel report with all profiling analysis data, quality metrics, drift detection, risk assessment, and readiness scores",
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "content_base64": base64.b64encode(file_content).decode('utf-8'),
                "size_bytes": len(file_content),
                "creation_date": datetime.utcnow().isoformat() + "Z",
                "sheets": ["Summary", "Profiler", "Drift Detection", "Risk Assessment", "Readiness", "Governance", "Test Coverage", "Alerts", "Issues", "Recommendations"]
            }
        except Exception as e:
            print(f"Error generating Excel report: {str(e)}")
            return {
                "download_id": f"{analysis_id}_profile_excel_error",
                "status": "error",
                "error": str(e)
            }
    
    def _create_analysis_summary_sheet(self, wb, analysis_id, execution_time_ms, alerts, issues, recommendations, executive_summary):
        """Create analysis summary sheet."""
        ws = wb.create_sheet("Summary", 0)
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 50
        
        row = 1
        # Title
        ws[f'A{row}'] = "PROFILE MY DATA - ANALYSIS SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 2
        
        # Metadata
        metadata = [
            ["Analysis ID", analysis_id],
            ["Tool", "Profile My Data"],
            ["Timestamp", datetime.utcnow().isoformat() + 'Z'],
            ["Execution Time (ms)", execution_time_ms],
            ["Report Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Executive Summary
        if executive_summary:
            ws[f'A{row}'] = "EXECUTIVE SUMMARY"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            ws[f'A{row}'].fill = self.subheader_fill
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            for item in executive_summary:
                ws[f'A{row}'] = item.get('title', '')
                ws[f'B{row}'] = f"{item.get('value', '')} - {item.get('description', '')}"
                ws[f'A{row}'].border = self.border
                ws[f'B{row}'].border = self.border
                row += 1
            
            row += 1
        
        # Statistics
        ws[f'A{row}'] = "ANALYSIS STATISTICS"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        stats = [
            ["Total Alerts", len(alerts)],
            ["Total Issues", len(issues)],
            ["Total Recommendations", len(recommendations)],
            ["High Severity Alerts", len([a for a in alerts if a.get('severity') == 'high'])],
            ["Critical Alerts", len([a for a in alerts if a.get('severity') == 'critical'])]
        ]
        
        for key, value in stats:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
    
    def _create_profiler_sheet(self, wb, agent_output):
        """Create unified profiler detailed sheet."""
        ws = wb.create_sheet("Profiler", 1)
        self._set_column_widths(ws, [25, 15, 12, 12, 12, 12, 50])
        
        row = 1
        ws[f'A{row}'] = "UNIFIED PROFILER ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Total Columns", summary_metrics.get("total_columns", 0)],
            ["Records Analyzed", summary_metrics.get("records_analyzed", 0)],
            ["Quality Grade", data.get("quality_summary", {}).get("overall_quality_grade", "")]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Quality scores
        ws[f'A{row}'] = "QUALITY SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        quality_summary = data.get("quality_summary", {})
        score_items = [
            ["Overall Quality Score", quality_summary.get("overall_quality_score", 0)],
            ["Completeness", quality_summary.get("completeness_score", 0)],
            ["Uniqueness", quality_summary.get("uniqueness_score", 0)],
            ["Validity", quality_summary.get("validity_score", 0)],
            ["Consistency", quality_summary.get("consistency_score", 0)]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Field analysis
        ws[f'A{row}'] = "FIELD QUALITY ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        headers = ["Field Name", "Data Type", "Quality Score", "Null %", "Completeness", "Uniqueness", "Distinct Values"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        for field in data.get("fields", []):
            ws.cell(row=row, column=1, value=field.get("field_name", ""))
            ws.cell(row=row, column=2, value=field.get("data_type", ""))
            ws.cell(row=row, column=3, value=field.get("quality_score", 0))
            
            props = field.get("properties", {})
            ws.cell(row=row, column=4, value=f"{props.get('null_percentage', 0):.2f}%")
            ws.cell(row=row, column=5, value=props.get("completeness_score", 0))
            ws.cell(row=row, column=6, value=props.get("uniqueness_score", 0))
            ws.cell(row=row, column=7, value=props.get("distinct_count", 0))
            
            for col_idx in range(1, 8):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _create_drift_sheet(self, wb, agent_output):
        """Create drift detector detailed sheet."""
        ws = wb.create_sheet("Drift Detection", 2)
        self._set_column_widths(ws, [25, 15, 12, 15, 15, 50])
        
        row = 1
        ws[f'A{row}'] = "DRIFT DETECTOR ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Dataset Stability", data.get("drift_summary", {}).get("dataset_stability", "")],
            ["Fields with Drift", data.get("drift_summary", {}).get("fields_with_drift", 0)],
            ["Drift Percentage", f"{data.get('drift_summary', {}).get('drift_percentage', 0):.2f}%"]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Drift summary
        ws[f'A{row}'] = "DRIFT SUMMARY METRICS"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        drift_summary = data.get("drift_summary", {})
        summary_items = [
            ["Overall Drift Score", drift_summary.get("overall_drift_score", 0)],
            ["Average PSI", drift_summary.get("average_psi", 0)],
            ["Max PSI", drift_summary.get("max_psi", 0)],
            ["Fields Analyzed", drift_summary.get("fields_analyzed", 0)]
        ]
        
        for key, value in summary_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Field drift analysis
        ws[f'A{row}'] = "FIELD DRIFT ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        headers = ["Field Name", "Drift Detected", "PSI Score", "KL Divergence", "Stability", "Action Required"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        for field in data.get("fields", []):
            drift_analysis = field.get("drift_analysis", {})
            ws.cell(row=row, column=1, value=field.get("field_name", ""))
            ws.cell(row=row, column=2, value="Yes" if drift_analysis.get("drift_detected", False) else "No")
            ws.cell(row=row, column=3, value=f"{drift_analysis.get('psi_score', 0):.4f}")
            ws.cell(row=row, column=4, value=f"{drift_analysis.get('kl_divergence', 0):.4f}")
            ws.cell(row=row, column=5, value=drift_analysis.get("stability_status", ""))
            ws.cell(row=row, column=6, value="Yes" if drift_analysis.get("drift_detected", False) else "No")
            
            for col_idx in range(1, 7):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _create_risk_sheet(self, wb, agent_output):
        """Create score risk detailed sheet."""
        ws = wb.create_sheet("Risk Assessment", 3)
        self._set_column_widths(ws, [25, 12, 12, 15, 50])
        
        row = 1
        ws[f'A{row}'] = "SCORE RISK ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Overall Risk Level", data.get("risk_summary", {}).get("overall_risk_level", "")],
            ["PII Fields Detected", summary_metrics.get("pii_fields_detected", 0)],
            ["High Risk Fields", summary_metrics.get("fields_with_high_risk", 0)]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Risk scores
        ws[f'A{row}'] = "RISK ASSESSMENT SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        risk_summary = data.get("risk_summary", {})
        score_items = [
            ["Overall Risk Score", risk_summary.get("overall_risk_score", 0)],
            ["PII Risk", risk_summary.get("pii_risk_score", 0)],
            ["Compliance Risk", risk_summary.get("compliance_risk_score", 0)],
            ["Sensitivity Risk", risk_summary.get("sensitivity_risk_score", 0)]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Field risk analysis
        ws[f'A{row}'] = "FIELD RISK ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        headers = ["Field Name", "Risk Score", "Risk Level", "PII Detected", "Risk Factors"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        for field in data.get("fields", []):
            risk_factors = [f.get('factor', '') for f in field.get('risk_factors', [])]
            ws.cell(row=row, column=1, value=field.get("field_name", ""))
            ws.cell(row=row, column=2, value=field.get("risk_score", 0))
            ws.cell(row=row, column=3, value=field.get("risk_level", ""))
            ws.cell(row=row, column=4, value="Yes" if any('pii' in f.lower() for f in risk_factors) else "No")
            ws.cell(row=row, column=5, value="; ".join(risk_factors))
            
            for col_idx in range(1, 6):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _create_readiness_sheet(self, wb, agent_output):
        """Create readiness rater detailed sheet."""
        ws = wb.create_sheet("Readiness", 4)
        self._set_column_widths(ws, [35, 50])
        
        row = 1
        ws[f'A{row}'] = "READINESS RATER ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Readiness assessment
        ws[f'A{row}'] = "READINESS ASSESSMENT"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        assessment = data.get("readiness_assessment", {})
        assessment_items = [
            ["Overall Score", assessment.get("overall_score", 0)],
            ["Overall Status", assessment.get("overall_status", "")],
            ["Production Ready", "Yes" if assessment.get("overall_status") == "ready" else "No"]
        ]
        
        for key, value in assessment_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 2
        
        # Deductions
        ws[f'A{row}'] = "READINESS DEDUCTIONS"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        headers = ["Deduction Reason", "Details"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        row += 1
        
        for deduction in data.get("deductions", []):
            ws.cell(row=row, column=1, value=deduction.get("deduction_reason", ""))
            ws.cell(row=row, column=2, value=f"Fields: {', '.join(deduction.get('fields_affected', []))}")
            ws.cell(row=row, column=1).border = self.border
            ws.cell(row=row, column=2).border = self.border
            ws.cell(row=row, column=2).alignment = self.left_alignment
            row += 1
    
    def _create_governance_sheet(self, wb, agent_output):
        """Create governance checker detailed sheet."""
        ws = wb.create_sheet("Governance", 5)
        self._set_column_widths(ws, [30, 15, 50])
        
        row = 1
        ws[f'A{row}'] = "GOVERNANCE CHECKER ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Compliance Status", data.get("compliance_status", "")],
            ["Governance Issues", summary_metrics.get("governance_issues_found", 0)]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Governance scores
        ws[f'A{row}'] = "GOVERNANCE SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        governance_scores = data.get("governance_scores", {})
        score_items = [
            ["Overall Score", governance_scores.get("overall", 0)],
            ["Lineage", governance_scores.get("lineage", 0)],
            ["Classification", governance_scores.get("classification", 0)],
            ["Consent Management", governance_scores.get("consent", 0)]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Governance issues
        ws[f'A{row}'] = "GOVERNANCE ISSUES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        headers = ["Issue Type", "Severity", "Message"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        row += 1
        
        for issue in data.get("governance_issues", []):
            ws.cell(row=row, column=1, value=issue.get("type", ""))
            ws.cell(row=row, column=2, value=issue.get("severity", ""))
            ws.cell(row=row, column=3, value=issue.get("message", ""))
            
            for col_idx in range(1, 4):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _create_test_coverage_sheet(self, wb, agent_output):
        """Create test coverage detailed sheet."""
        ws = wb.create_sheet("Test Coverage", 6)
        self._set_column_widths(ws, [30, 15, 50])
        
        row = 1
        ws[f'A{row}'] = "TEST COVERAGE ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Coverage Status", data.get("coverage_status", "")],
            ["Test Issues", summary_metrics.get("test_issues_count", 0)]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Test scores
        ws[f'A{row}'] = "TEST COVERAGE SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        test_scores = data.get("test_coverage_scores", {})
        score_items = [
            ["Overall Score", test_scores.get("overall", 0)],
            ["Uniqueness", test_scores.get("uniqueness", 0)],
            ["Range", test_scores.get("range", 0)],
            ["Format", test_scores.get("format", 0)]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
    
    def _create_alerts_sheet(self, wb, alerts):
        """Create alerts sheet."""
        ws = wb.create_sheet("Alerts", 7)
        self._set_column_widths(ws, [20, 15, 20, 50, 20, 30])
        
        row = 1
        headers = ["Alert ID", "Severity", "Category", "Message", "Affected Fields", "Recommendation"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        for alert in alerts:
            ws.cell(row=row, column=1, value=alert.get("alert_id", ""))
            ws.cell(row=row, column=2, value=alert.get("severity", ""))
            ws.cell(row=row, column=3, value=alert.get("category", ""))
            ws.cell(row=row, column=4, value=alert.get("message", ""))
            ws.cell(row=row, column=5, value=alert.get("affected_fields_count", ""))
            ws.cell(row=row, column=6, value=alert.get("recommendation", ""))
            
            for col_idx in range(1, 7):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _create_issues_sheet(self, wb, issues):
        """Create issues sheet."""
        ws = wb.create_sheet("Issues", 8)
        self._set_column_widths(ws, [20, 20, 20, 20, 15, 50])
        
        row = 1
        headers = ["Issue ID", "Agent", "Field", "Issue Type", "Severity", "Message"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        for issue in issues:
            ws.cell(row=row, column=1, value=issue.get("issue_id", ""))
            ws.cell(row=row, column=2, value=issue.get("agent_id", ""))
            ws.cell(row=row, column=3, value=issue.get("field_name", ""))
            ws.cell(row=row, column=4, value=issue.get("issue_type", ""))
            ws.cell(row=row, column=5, value=issue.get("severity", ""))
            ws.cell(row=row, column=6, value=issue.get("message", ""))
            
            for col_idx in range(1, 7):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _create_recommendations_sheet(self, wb, recommendations):
        """Create recommendations sheet."""
        ws = wb.create_sheet("Recommendations", 9)
        self._set_column_widths(ws, [25, 20, 20, 15, 50, 15])
        
        row = 1
        headers = ["Recommendation ID", "Agent", "Field", "Priority", "Recommendation", "Timeline"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        for rec in recommendations:
            ws.cell(row=row, column=1, value=rec.get("recommendation_id", ""))
            ws.cell(row=row, column=2, value=rec.get("agent_id", ""))
            ws.cell(row=row, column=3, value=rec.get("field_name", ""))
            ws.cell(row=row, column=4, value=rec.get("priority", ""))
            ws.cell(row=row, column=5, value=rec.get("recommendation", ""))
            ws.cell(row=row, column=6, value=rec.get("timeline", ""))
            
            for col_idx in range(1, 7):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _set_column_widths(self, ws, widths):
        """Set column widths for worksheet."""
        for col_idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    def _generate_json_report(
        self,
        agent_results: Dict[str, Any],
        analysis_id: str,
        execution_time_ms: int,
        alerts: List[Dict],
        issues: List[Dict],
        recommendations: List[Dict],
        executive_summary: List[Dict]
    ) -> Dict[str, Any]:
        """Generate comprehensive JSON report with all agent data."""
        try:
            report_data = {
                "metadata": {
                    "analysis_id": analysis_id,
                    "tool": "profile-my-data",
                    "timestamp": datetime.utcnow().isoformat() + "Z",
                    "execution_time_ms": execution_time_ms,
                    "report_version": "1.0"
                },
                "executive_summary": executive_summary,
                "summary": {
                    "total_alerts": len(alerts),
                    "total_issues": len(issues),
                    "total_recommendations": len(recommendations),
                    "critical_alerts": len([a for a in alerts if a.get('severity') == 'critical']),
                    "high_severity_alerts": len([a for a in alerts if a.get('severity') == 'high']),
                    "medium_severity_alerts": len([a for a in alerts if a.get('severity') == 'medium'])
                },
                "alerts": alerts,
                "issues": issues,
                "recommendations": recommendations,
                "agent_results": {}
            }
            
            # Add complete agent outputs
            for agent_id, agent_output in agent_results.items():
                if agent_output.get("status") == "success":
                    report_data["agent_results"][agent_id] = {
                        "status": agent_output.get("status"),
                        "execution_time_ms": agent_output.get("execution_time_ms", 0),
                        "summary_metrics": agent_output.get("summary_metrics", {}),
                        "data": agent_output.get("data", {}),
                        "timestamp": agent_output.get("timestamp", "")
                    }
            
            # Convert to JSON string
            json_str = json.dumps(report_data, indent=2, default=str)
            json_bytes = json_str.encode('utf-8')
            
            return {
                "download_id": f"{analysis_id}_profile_json",
                "name": "Profile My Data - Complete Analysis JSON",
                "format": "json",
                "file_name": "profile_my_data_analysis.json",
                "description": "Complete hierarchical JSON report with all analysis data, including raw agent outputs, quality metrics, drift detection, risk assessment, and readiness scores",
                "mimeType": "application/json",
                "content_base64": base64.b64encode(json_bytes).decode('utf-8'),
                "size_bytes": len(json_bytes),
                "creation_date": datetime.utcnow().isoformat() + "Z"
            }
        except Exception as e:
            print(f"Error generating JSON report: {str(e)}")
            return {
                "download_id": f"{analysis_id}_profile_json_error",
                "status": "error",
                "error": str(e)
            }
