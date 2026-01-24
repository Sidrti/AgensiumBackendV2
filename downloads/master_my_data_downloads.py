"""
Master My Data Downloads Module (Refactored)

Uses BaseDownloader for standardized report generation.
"""

from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from downloads.downloads_utils import BaseDownloader, load_tool_config

class MasterMyDataDownloads(BaseDownloader):
    """Handles comprehensive downloads for master-my-data tool."""
    
    def __init__(self, tool_id: str, tool_display_name: str = None):
        if not tool_display_name:
            config = load_tool_config(tool_id)
            tool_display_name = config.get("tool", {}).get("name", tool_id)
        super().__init__(tool_id, tool_display_name)
        
    def create_tool_specific_sheets(self, wb: Workbook, agent_results: Dict[str, Any]):
        """Create tool-specific analysis sheets."""
        
        # 1. KEY IDENTIFIER SHEET
        key_identifier_output = agent_results.get("key-identifier", {})
        if key_identifier_output.get("status") == "success":
            self._create_key_identifier_sheet(wb, key_identifier_output)
        
        # 2. CONTRACT ENFORCER SHEET
        contract_enforcer_output = agent_results.get("contract-enforcer", {})
        if contract_enforcer_output.get("status") == "success":
            self._create_contract_enforcer_sheet(wb, contract_enforcer_output)
        
        # 3. SEMANTIC MAPPER SHEET
        semantic_mapper_output = agent_results.get("semantic-mapper", {})
        if semantic_mapper_output.get("status") == "success":
            self._create_semantic_mapper_sheet(wb, semantic_mapper_output)
        
        # 4. LINEAGE TRACER SHEET
        lineage_tracer_output = agent_results.get("lineage-tracer", {})
        if lineage_tracer_output.get("status") == "success":
            self._create_lineage_tracer_sheet(wb, lineage_tracer_output)
        
        # 5. GOLDEN RECORD BUILDER SHEET
        golden_record_output = agent_results.get("golden-record-builder", {})
        if golden_record_output.get("status") == "success":
            self._create_golden_record_sheet(wb, golden_record_output)
        
        # 6. SURVIVORSHIP RESOLVER SHEET
        survivorship_output = agent_results.get("survivorship-resolver", {})
        if survivorship_output.get("status") == "success":
            self._create_survivorship_sheet(wb, survivorship_output)
        
        # 7. MASTER WRITEBACK SHEET
        master_writeback_output = agent_results.get("master-writeback-agent", {})
        if master_writeback_output.get("status") == "success":
            self._create_master_writeback_sheet(wb, master_writeback_output)
        
        # 8. STEWARDSHIP FLAGGER SHEET
        stewardship_output = agent_results.get("stewardship-flagger", {})
        if stewardship_output.get("status") == "success":
            self._create_stewardship_sheet(wb, stewardship_output)

    def _create_key_identifier_sheet(self, wb, agent_output):
        """Create key identifier detailed sheet."""
        ws = wb.create_sheet("Key Identifier")
        self.styler.set_column_widths(ws, [25, 20, 15, 15, 50])
        
        row = 1
        ws[f'A{row}'] = "KEY IDENTIFIER ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        # Metadata section
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Columns Analyzed", summary_metrics.get("columns_analyzed", 0)],
            ["Keys Identified", summary_metrics.get("keys_identified", 0)],
            ["Composite Keys Found", summary_metrics.get("composite_keys_found", 0)],
            ["Candidate PKs", summary_metrics.get("candidate_pks", 0)],
            ["Candidate FKs", summary_metrics.get("candidate_fks", 0)]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Key Analysis Score
        ws[f'A{row}'] = "KEY ANALYSIS SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        key_score = data.get("key_score", {})
        score_items = [
            ["Overall Score", key_score.get("overall_score", 0)],
            ["Quality Status", key_score.get("quality_status", "unknown")],
            ["Key Coverage", f"{key_score.get('key_coverage', 0):.1f}%"],
            ["Uniqueness Score", key_score.get("uniqueness_score", 0)]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Identified Keys table
        ws[f'A{row}'] = "IDENTIFIED KEYS"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        headers = ["Column(s)", "Key Type", "Confidence", "Uniqueness %", "Notes"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        key_analysis = data.get("key_analysis", {})
        identified_keys = key_analysis.get("identified_keys", [])
        
        for key_info in identified_keys:
            columns = key_info.get("columns", [])
            if isinstance(columns, list):
                col_str = ", ".join(columns)
            else:
                col_str = str(columns)
            
            ws.cell(row=row, column=1, value=col_str)
            ws.cell(row=row, column=2, value=key_info.get("key_type", "").upper())
            ws.cell(row=row, column=3, value=f"{key_info.get('confidence', 0):.2f}")
            ws.cell(row=row, column=4, value=f"{key_info.get('uniqueness_percentage', 0):.2f}%")
            ws.cell(row=row, column=5, value=key_info.get("notes", ""))
            
            for col_idx in range(1, 6):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
        
        row += 1
        
        # Key Candidates table
        ws[f'A{row}'] = "KEY CANDIDATES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        headers = ["Column", "Candidate Type", "Uniqueness %", "Null %", "Recommendation"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        key_candidates = key_analysis.get("key_candidates", [])
        
        for candidate in key_candidates:
            ws.cell(row=row, column=1, value=candidate.get("column", ""))
            ws.cell(row=row, column=2, value=candidate.get("candidate_type", ""))
            ws.cell(row=row, column=3, value=f"{candidate.get('uniqueness_percentage', 0):.2f}%")
            ws.cell(row=row, column=4, value=f"{candidate.get('null_percentage', 0):.2f}%")
            ws.cell(row=row, column=5, value=candidate.get("recommendation", ""))
            
            for col_idx in range(1, 6):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _create_contract_enforcer_sheet(self, wb, agent_output):
        """Create contract enforcer detailed sheet."""
        ws = wb.create_sheet("Contract Enforcer")
        self.styler.set_column_widths(ws, [25, 20, 15, 15, 50])
        
        row = 1
        ws[f'A{row}'] = "CONTRACT ENFORCER ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        # Metadata section
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Total Records", summary_metrics.get("total_records", 0)],
            ["Compliant Records", summary_metrics.get("compliant_records", 0)],
            ["Non-Compliant Records", summary_metrics.get("non_compliant_records", 0)],
            ["Compliance Score", f"{summary_metrics.get('compliance_score', 0):.1f}%"],
            ["Auto-Transformations", summary_metrics.get("auto_transformations_applied", 0)]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Compliance Score
        ws[f'A{row}'] = "COMPLIANCE SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        contract_score = data.get("contract_score", {})
        metrics = contract_score.get("metrics", {})
        score_items = [
            ["Overall Score", contract_score.get("overall_score", 0)],
            ["Quality Status", contract_score.get("quality_status", "unknown")],
            ["Structural Compliance", f"{metrics.get('structural_compliance', 0):.1f}%"],
            ["Value Compliance", f"{metrics.get('value_compliance', 0):.1f}%"],
            ["Type Compliance", f"{metrics.get('type_compliance', 0):.1f}%"]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Violations table
        ws[f'A{row}'] = "CONTRACT VIOLATIONS"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        headers = ["Column", "Violation Type", "Severity", "Count", "Details"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        contract_analysis = data.get("contract_analysis", {})
        violations = contract_analysis.get("violations", [])
        
        for violation in violations:
            ws.cell(row=row, column=1, value=violation.get("column", ""))
            ws.cell(row=row, column=2, value=violation.get("violation_type", ""))
            ws.cell(row=row, column=3, value=violation.get("severity", "").upper())
            ws.cell(row=row, column=4, value=violation.get("count", 0))
            ws.cell(row=row, column=5, value=violation.get("details", ""))
            
            # Color code severity
            severity_cell = ws.cell(row=row, column=3)
            severity = violation.get("severity", "").lower()
            if severity == "critical":
                severity_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            elif severity == "high":
                severity_cell.fill = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
            elif severity == "medium":
                severity_cell.fill = PatternFill(start_color="FFEB99", end_color="FFEB99", fill_type="solid")
            
            for col_idx in range(1, 6):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
        
        row += 1
        
        # Field Compliance table
        ws[f'A{row}'] = "FIELD COMPLIANCE"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        headers = ["Field", "Expected Type", "Actual Type", "Compliant", "Notes"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        field_compliance = contract_analysis.get("field_compliance", {})
        
        for field_name, field_data in field_compliance.items():
            ws.cell(row=row, column=1, value=field_name)
            ws.cell(row=row, column=2, value=field_data.get("expected_type", ""))
            ws.cell(row=row, column=3, value=field_data.get("actual_type", ""))
            
            compliant = field_data.get("compliant", False)
            ws.cell(row=row, column=4, value="YES" if compliant else "NO")
            compliance_cell = ws.cell(row=row, column=4)
            if compliant:
                compliance_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            else:
                compliance_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            ws.cell(row=row, column=5, value=field_data.get("notes", ""))
            
            for col_idx in range(1, 6):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _create_semantic_mapper_sheet(self, wb, agent_output):
        """Create semantic mapper detailed sheet."""
        ws = wb.create_sheet("Semantic Mapper")
        self.styler.set_column_widths(ws, [25, 25, 15, 20, 50])
        
        row = 1
        ws[f'A{row}'] = "SEMANTIC MAPPER ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        # Metadata section
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Total Columns", summary_metrics.get("total_columns", 0)],
            ["Mapped Columns", summary_metrics.get("mapped_columns", 0)],
            ["Unmapped Columns", summary_metrics.get("unmapped_columns", 0)],
            ["High Confidence Mappings", summary_metrics.get("high_confidence_mappings", 0)],
            ["Average Confidence", f"{summary_metrics.get('average_confidence', 0):.2f}"],
            ["Total Transformations", summary_metrics.get("total_transformations", 0)]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Mapping Score
        ws[f'A{row}'] = "MAPPING SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        mapping_score = data.get("mapping_score", 0)
        if isinstance(mapping_score, dict):
            overall_score = mapping_score.get("overall_score", 0)
        else:
            overall_score = mapping_score
        
        quality_status = data.get("quality_status", "unknown")
        statistics = data.get("statistics", {})
        
        score_items = [
            ["Overall Score", overall_score],
            ["Quality Status", quality_status],
            ["Mapped Columns", f"{statistics.get('mapped_columns', 0)}/{statistics.get('total_columns', 0)}"],
            ["Average Confidence", f"{statistics.get('average_confidence', 0):.2f}"]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Column Mappings table
        ws[f'A{row}'] = "COLUMN MAPPINGS"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        headers = ["Source Column", "Standard Name", "Confidence", "Semantic Type", "Status"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        column_mappings = data.get("column_mappings", [])
        
        for mapping in column_mappings[:50]:  # Limit to 50
            ws.cell(row=row, column=1, value=mapping.get("original_name", ""))
            ws.cell(row=row, column=2, value=mapping.get("standard_name", ""))
            ws.cell(row=row, column=3, value=f"{mapping.get('confidence', 0):.2f}")
            ws.cell(row=row, column=4, value=mapping.get("semantic_type", ""))
            ws.cell(row=row, column=5, value=mapping.get("status", ""))
            
            for col_idx in range(1, 6):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
        
        row += 1
        
        # Unmapped Columns
        unmapped_columns = data.get("unmapped_columns", [])
        if unmapped_columns:
            ws[f'A{row}'] = "UNMAPPED COLUMNS"
            ws[f'A{row}'].font = Font(bold=True, size=10)
            ws[f'A{row}'].fill = self.subheader_fill
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            headers = ["Column", "Reason", "Suggested Action"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.border = self.border
                cell.alignment = self.center_alignment
            row += 1
            
            for unmapped in unmapped_columns[:20]:
                ws.cell(row=row, column=1, value=unmapped.get("column", ""))
                ws.cell(row=row, column=2, value=unmapped.get("reason", "No match found"))
                ws.cell(row=row, column=3, value="Add custom mapping")
                
                for col_idx in range(1, 4):
                    ws.cell(row=row, column=col_idx).border = self.border
                    ws.cell(row=row, column=col_idx).alignment = self.left_alignment
                row += 1
    
    def _create_lineage_tracer_sheet(self, wb, agent_output):
        """Create lineage tracer detailed sheet."""
        ws = wb.create_sheet("Lineage Tracer")
        self.styler.set_column_widths(ws, [25, 20, 20, 50])
        
        row = 1
        ws[f'A{row}'] = "LINEAGE TRACER ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        # Metadata section
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Sources Identified", summary_metrics.get("sources_identified", 0)],
            ["Transformations Tracked", summary_metrics.get("transformations_tracked", 0)],
            ["Max Depth", summary_metrics.get("max_depth", 0)]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Transformation Log
        ws[f'A{row}'] = "TRANSFORMATION LOG"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        headers = ["Step", "Transformation", "Source", "Details"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        lineage_analysis = data.get("lineage_analysis", {})
        transformation_log = lineage_analysis.get("transformation_log", [])
        
        for idx, transform in enumerate(transformation_log, 1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=transform.get("transformation", ""))
            ws.cell(row=row, column=3, value=transform.get("source", ""))
            ws.cell(row=row, column=4, value=transform.get("details", ""))
            
            for col_idx in range(1, 5):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _create_golden_record_sheet(self, wb, agent_output):
        """Create golden record builder detailed sheet."""
        ws = wb.create_sheet("Golden Records")
        self.styler.set_column_widths(ws, [20, 15, 15, 15, 50])
        
        row = 1
        ws[f'A{row}'] = "GOLDEN RECORD BUILDER ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        # Metadata section
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Input Records", summary_metrics.get("input_records", 0)],
            ["Golden Records Created", summary_metrics.get("golden_records_created", 0)],
            ["Clusters Formed", summary_metrics.get("clusters_formed", 0)],
            ["Compression Ratio", f"{summary_metrics.get('compression_ratio', 0):.2f}x"],
            ["Conflicts Resolved", summary_metrics.get("conflicts_resolved", 0)],
            ["Average Trust Score", f"{summary_metrics.get('average_trust_score', 0):.3f}"]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Golden Record Score
        ws[f'A{row}'] = "GOLDEN RECORD SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        # Get score directly as float
        golden_score = data.get("golden_score", 0)
        if isinstance(golden_score, dict):
            overall_score = golden_score.get("overall_score", 0)
        else:
            overall_score = golden_score
        
        quality_status = data.get("quality_status", "unknown")
        statistics = data.get("statistics", {})
        
        score_items = [
            ["Overall Score", overall_score],
            ["Quality Status", quality_status],
            ["Compression Ratio", f"{statistics.get('compression_ratio', 0):.2f}x"],
            ["Average Trust Score", f"{statistics.get('average_trust_score', 0):.3f}"]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Survivorship Rules Applied
        ws[f'A{row}'] = "SURVIVORSHIP RULES APPLIED"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        headers = ["Column/Rule", "Value"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        survivorship_rules = data.get("survivorship_rules_applied", {})
        for rule_name, rule_value in survivorship_rules.items():
            ws.cell(row=row, column=1, value=rule_name)
            ws.cell(row=row, column=2, value=str(rule_value))
            for col_idx in range(1, 3):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
        
        row += 1
        
        # Field Resolutions
        ws[f'A{row}'] = "FIELD RESOLUTIONS"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        headers = ["Cluster", "Column", "Resolution Method", "Winner Value", "Trust Score"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        field_resolutions = data.get("field_resolutions", [])
        
        for resolution in field_resolutions[:50]:  # Limit to 50 rows
            ws.cell(row=row, column=1, value=resolution.get("cluster_id", ""))
            ws.cell(row=row, column=2, value=resolution.get("column", ""))
            ws.cell(row=row, column=3, value=resolution.get("resolution_method", ""))
            winner = resolution.get("winner_value", "")
            ws.cell(row=row, column=4, value=str(winner)[:50] if winner else "")
            ws.cell(row=row, column=5, value=f"{resolution.get('trust_score', 0):.2f}")
            
            for col_idx in range(1, 6):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _create_survivorship_sheet(self, wb, agent_output):
        """Create survivorship resolver detailed sheet."""
        ws = wb.create_sheet("Survivorship")
        self.styler.set_column_widths(ws, [25, 20, 15, 15, 50])
        
        row = 1
        ws[f'A{row}'] = "SURVIVORSHIP RESOLVER ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        # Metadata section
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Conflicts Detected", summary_metrics.get("conflicts_detected", 0)],
            ["Conflicts Resolved", summary_metrics.get("conflicts_resolved", 0)],
            ["Values Survived", summary_metrics.get("values_survived", 0)],
            ["Unresolved Conflicts", summary_metrics.get("unresolved_conflicts", 0)],
            ["Resolution Rate", f"{summary_metrics.get('resolution_rate', 0):.1f}%"],
            ["Average Confidence", f"{summary_metrics.get('average_confidence', 0):.2f}"]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Survivorship Score
        ws[f'A{row}'] = "SURVIVORSHIP SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        # Get score directly as float
        survivorship_score = data.get("survivorship_score", 0)
        if isinstance(survivorship_score, dict):
            overall_score = survivorship_score.get("overall_score", 0)
        else:
            overall_score = survivorship_score
        
        quality_status = data.get("quality_status", "unknown")
        statistics = data.get("statistics", {})
        
        score_items = [
            ["Overall Score", overall_score],
            ["Quality Status", quality_status],
            ["Resolution Rate", f"{statistics.get('resolution_rate', 0):.1f}%"],
            ["Average Confidence", f"{statistics.get('average_confidence', 0):.2f}"]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Rules Applied
        ws[f'A{row}'] = "RULES APPLIED"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        headers = ["Rule", "Count"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        rules_applied = data.get("rules_applied", {})
        for rule_name, count in rules_applied.items():
            ws.cell(row=row, column=1, value=rule_name)
            ws.cell(row=row, column=2, value=count)
            for col_idx in range(1, 3):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
        
        row += 1
        
        # Resolution Log
        ws[f'A{row}'] = "RESOLUTION LOG"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        headers = ["Cluster", "Column", "Rule Applied", "Confidence", "Rationale"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        resolution_log = data.get("resolution_log", [])
        
        for resolution in resolution_log[:50]:  # Limit to 50 rows
            ws.cell(row=row, column=1, value=resolution.get("cluster_id", ""))
            ws.cell(row=row, column=2, value=resolution.get("column", ""))
            ws.cell(row=row, column=3, value=resolution.get("rule", ""))
            ws.cell(row=row, column=4, value=f"{resolution.get('confidence', 0):.2f}")
            ws.cell(row=row, column=5, value=resolution.get("rationale", ""))
            
            for col_idx in range(1, 6):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
    
    def _create_master_writeback_sheet(self, wb, agent_output):
        """Create master writeback detailed sheet."""
        ws = wb.create_sheet("Master Writeback")
        self.styler.set_column_widths(ws, [30, 25, 50])
        
        row = 1
        ws[f'A{row}'] = "MASTER WRITEBACK ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        # Metadata section
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Records Written", summary_metrics.get("records_written", 0)],
            ["Bytes Written", summary_metrics.get("bytes_written", 0)],
            ["Versions Created", summary_metrics.get("versions_created", 0)]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Writeback Score
        ws[f'A{row}'] = "WRITEBACK SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        writeback_score = data.get("writeback_score", {})
        metrics = writeback_score.get("metrics", {})
        score_items = [
            ["Overall Score", writeback_score.get("overall_score", 0)],
            ["Quality Status", writeback_score.get("quality_status", "unknown")],
            ["Data Integrity", f"{metrics.get('data_integrity', 0):.1f}%"],
            ["Write Success Rate", f"{metrics.get('write_success_rate', 0):.1f}%"]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Output Details
        ws[f'A{row}'] = "OUTPUT DETAILS"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        writeback_analysis = data.get("writeback_analysis", {})
        output_details = [
            ["Output Location", writeback_analysis.get("output_location", "N/A")],
            ["Audit Log Location", writeback_analysis.get("audit_log_location", "N/A")],
            ["Format", writeback_analysis.get("format", "N/A")],
            ["Versioning Enabled", "Yes" if writeback_analysis.get("versioning_enabled", False) else "No"]
        ]
        
        for key, value in output_details:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
    
    def _create_stewardship_sheet(self, wb, agent_output):
        """Create stewardship flagger detailed sheet."""
        ws = wb.create_sheet("Stewardship")
        self.styler.set_column_widths(ws, [15, 20, 15, 15, 50])
        
        row = 1
        ws[f'A{row}'] = "STEWARDSHIP FLAGGER ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        
        # Metadata section
        metadata = [
            ["Status", agent_output.get("status")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Total Records", summary_metrics.get("total_records", 0)],
            ["Total Issues", summary_metrics.get("total_issues", 0)],
            ["Tasks Created", summary_metrics.get("tasks_created", 0)],
            ["Critical Issues", summary_metrics.get("critical_issues", 0)],
            ["High Priority Tasks", summary_metrics.get("high_priority_tasks", 0)],
            ["Records Flagged", summary_metrics.get("records_flagged", 0)],
            ["Clean Data Rate", f"{summary_metrics.get('clean_data_rate', 0):.1f}%"]
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Stewardship Score
        ws[f'A{row}'] = "STEWARDSHIP SCORES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        row += 1
        
        # Get score directly as float
        stewardship_score = data.get("stewardship_score", 0)
        if isinstance(stewardship_score, dict):
            overall_score = stewardship_score.get("overall_score", 0)
        else:
            overall_score = stewardship_score
        
        quality_status = data.get("quality_status", "unknown")
        statistics = data.get("statistics", {})
        
        score_items = [
            ["Overall Score", overall_score],
            ["Quality Status", quality_status],
            ["Clean Data Rate", f"{statistics.get('clean_data_rate', 0):.1f}%"],
            ["Records Flagged", statistics.get("records_flagged", 0)]
        ]
        
        for key, value in score_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Issue Distribution
        ws[f'A{row}'] = "ISSUE DISTRIBUTION"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        headers = ["Issue Type", "Count", "Percentage"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        issue_distribution = data.get("issue_distribution", {})
        total_issues = sum(issue_distribution.values()) if issue_distribution else 1
        
        for issue_type, count in issue_distribution.items():
            ws.cell(row=row, column=1, value=issue_type.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{(count / max(total_issues, 1) * 100):.1f}%")
            
            for col_idx in range(1, 4):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
        
        row += 1
        
        # Issues by Severity
        ws[f'A{row}'] = "ISSUES BY SEVERITY"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        headers = ["Severity", "Count"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        issues_by_severity = data.get("issues_by_severity", {})
        severity_order = ["critical", "high", "medium", "low", "info"]
        
        for severity in severity_order:
            count = issues_by_severity.get(severity, 0)
            if count > 0:
                ws.cell(row=row, column=1, value=severity.upper())
                ws.cell(row=row, column=2, value=count)
                
                # Color code severity
                severity_cell = ws.cell(row=row, column=1)
                if severity == "critical":
                    severity_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                elif severity == "high":
                    severity_cell.fill = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
                elif severity == "medium":
                    severity_cell.fill = PatternFill(start_color="FFEB99", end_color="FFEB99", fill_type="solid")
                
                for col_idx in range(1, 3):
                    ws.cell(row=row, column=col_idx).border = self.border
                    ws.cell(row=row, column=col_idx).alignment = self.left_alignment
                row += 1
        
        row += 1
        
        # Tasks List
        ws[f'A{row}'] = "STEWARDSHIP TASKS"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        headers = ["Entity ID", "Field", "Issue Type", "Priority", "Recommended Action"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
        row += 1
        
        task_list = data.get("task_list", [])
        
        for task in task_list[:100]:  # Limit to 100 rows
            ws.cell(row=row, column=1, value=str(task.get("entity_id", ""))[:20])
            ws.cell(row=row, column=2, value=task.get("field", ""))
            ws.cell(row=row, column=3, value=task.get("issue_type", ""))
            ws.cell(row=row, column=4, value=str(task.get("priority", "")).upper())
            ws.cell(row=row, column=5, value=str(task.get("recommended_action", ""))[:50])
            
            # Color code priority
            priority_cell = ws.cell(row=row, column=4)
            priority = str(task.get("priority", "")).lower()
            if priority == "critical":
                priority_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            elif priority == "high":
                priority_cell.fill = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
            elif priority == "medium":
                priority_cell.fill = PatternFill(start_color="FFEB99", end_color="FFEB99", fill_type="solid")
            
            for col_idx in range(1, 6):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1