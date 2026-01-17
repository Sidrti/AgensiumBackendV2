"""
Analyze My Data Downloads Module

Uses BaseDownloader for standardized report generation.
Handles downloads for analytics agents: Customer Segmentation, Market Basket, Experimental Design, Synthetic Control.
"""

from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from downloads.downloads_utils import BaseDownloader


class AnalyzeMyDataDownloads(BaseDownloader):
    """Handles comprehensive downloads for analyze-my-data tool."""
    
    def __init__(self):
        super().__init__("analyze-my-data", "Analyze My Data")
        
    def create_tool_specific_sheets(self, wb: Workbook, agent_results: Dict[str, Any]):
        """Create tool-specific analysis sheets."""
        
        # 1. CUSTOMER SEGMENTATION SHEET
        segmentation_output = agent_results.get("customer-segmentation-agent", {})
        if segmentation_output.get("status") == "success":
            self._create_segmentation_sheet(wb, segmentation_output)
            self._create_segment_customers_sheet(wb, segmentation_output)
        
        # 2. MARKET BASKET & SEQUENCE SHEET
        market_output = agent_results.get("market-basket-sequence-agent", {})
        if market_output.get("status") == "success":
            self._create_market_basket_sheet(wb, market_output)
        
        # 3. EXPERIMENTAL DESIGN SHEET
        experiment_output = agent_results.get("experimental-design-agent", {})
        if experiment_output.get("status") == "success":
            self._create_experiment_design_sheet(wb, experiment_output)
        
        # 4. SYNTHETIC CONTROL SHEET
        synthetic_output = agent_results.get("synthetic-control-agent", {})
        if synthetic_output.get("status") == "success":
            self._create_synthetic_control_sheet(wb, synthetic_output)
            self._create_synthetic_time_series_sheet(wb, synthetic_output)

    def _create_segmentation_sheet(self, wb: Workbook, agent_output: Dict[str, Any]):
        """Create customer segmentation summary sheet."""
        ws = wb.create_sheet("Customer Segmentation")
        self.styler.set_column_widths(ws, [25, 20, 18, 18, 18, 18, 20])
        
        row = 1
        
        # Page Title
        ws[f'A{row}'] = "CUSTOMER SEGMENTATION ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:G{row}')
        row += 2
        
        # Extract data
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        segmentation = data.get("segmentation", {})
        quality = data.get("quality", {})
        
        # Metadata Section
        metadata = [
            ["Status", agent_output.get("status", "")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Segmentation Mode", segmentation.get("mode", "").upper()],
            ["Timeframe", segmentation.get("timeframe", "")],
            ["Analysis Start Date", segmentation.get("analysis_window", {}).get("start_date", "")],
            ["Analysis End Date", segmentation.get("analysis_window", {}).get("end_date", "")],
            ["Total Customers", summary_metrics.get("total_customers", 0)],
            ["Segments Produced", summary_metrics.get("segments_produced", 0)],
            ["Total Rows Processed", summary_metrics.get("total_rows_processed", 0)],
            ["Invalid Rows", summary_metrics.get("invalid_rows", 0)],
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value) if value is not None else ""
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Quality Section
        ws[f'A{row}'] = "DATA QUALITY"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        quality_items = [
            ["Input Quality Score", quality.get("input_quality_score", 0)],
            ["Quality Status", quality.get("quality_status", "")],
            ["Invalid Rate (%)", f"{quality.get('invalid_rate_pct', 0):.2f}%"],
            ["Valid Rows", quality.get("valid_rows", 0)],
            ["Rows After Timeframe", quality.get("rows_after_timeframe", 0)],
            ["Timeframe Retained (%)", f"{quality.get('timeframe_retained_rate_pct', 0):.2f}%"],
        ]
        
        for key, value in quality_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value) if value is not None else ""
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Columns Used Section
        columns_used = segmentation.get("columns_used", {})
        ws[f'A{row}'] = "COLUMNS USED"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        columns_info = [
            ["Customer ID Column", columns_used.get("customer_id_column", "")],
            ["Transaction Date Column", columns_used.get("transaction_date_column", "")],
            ["Value Column", columns_used.get("value_column", "")],
        ]
        
        for key, value in columns_info:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value) if value is not None else ""
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 2
        
        # Segment Summary Table
        ws[f'A{row}'] = "SEGMENT SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        headers = ["Segment ID", "Label", "Description", "Customer Count", "Total Value", "Avg Value", "Value Share (%)"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.border
        row += 1
        
        # Segment rows
        for segment in data.get("segment_summary", []):
            ws.cell(row=row, column=1, value=segment.get("segment_id", ""))
            ws.cell(row=row, column=2, value=segment.get("segment_label", ""))
            ws.cell(row=row, column=3, value=segment.get("segment_description", ""))
            ws.cell(row=row, column=4, value=segment.get("customer_count", 0))
            
            total_value = segment.get("total_value", 0)
            ws.cell(row=row, column=5, value=f"{total_value:,.2f}" if total_value else "0.00")
            
            avg_value = segment.get("avg_value", 0)
            ws.cell(row=row, column=6, value=f"{avg_value:,.2f}" if avg_value else "0.00")
            
            value_share = segment.get("value_share_pct", 0)
            ws.cell(row=row, column=7, value=f"{value_share:.2f}%" if value_share else "0.00%")
            
            for col_idx in range(1, 8):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
        
        row += 2
        
        # Additional Segment Metrics Table
        ws[f'A{row}'] = "SEGMENT BEHAVIOR METRICS"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        behavior_headers = ["Segment Label", "Avg Frequency", "Avg Recency (Days)", "Customer %"]
        for col_idx, header in enumerate(behavior_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.border
        row += 1
        
        total_customers = summary_metrics.get("total_customers", 1) or 1
        for segment in data.get("segment_summary", []):
            ws.cell(row=row, column=1, value=segment.get("segment_label", ""))
            
            avg_freq = segment.get("avg_frequency", 0)
            ws.cell(row=row, column=2, value=f"{avg_freq:.2f}" if avg_freq else "0.00")
            
            avg_recency = segment.get("avg_recency_days", 0)
            ws.cell(row=row, column=3, value=f"{avg_recency:.1f}" if avg_recency else "0.0")
            
            customer_count = segment.get("customer_count", 0)
            customer_pct = (customer_count / total_customers) * 100
            ws.cell(row=row, column=4, value=f"{customer_pct:.2f}%")
            
            for col_idx in range(1, 5):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1

    def _create_segment_customers_sheet(self, wb: Workbook, agent_output: Dict[str, Any]):
        """Create detailed customer segments sample sheet."""
        ws = wb.create_sheet("Customer Details")
        self.styler.set_column_widths(ws, [20, 12, 20, 12, 18, 15, 18, 20])
        
        row = 1
        
        # Page Title
        ws[f'A{row}'] = "CUSTOMER SEGMENT ASSIGNMENTS (SAMPLE)"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:H{row}')
        row += 2
        
        data = agent_output.get("data", {})
        customer_sample = data.get("customer_segments_sample", [])
        
        ws[f'A{row}'] = f"Showing top {len(customer_sample)} customers by segment and monetary value"
        ws[f'A{row}'].font = Font(italic=True)
        row += 2
        
        # Table Headers
        headers = ["Customer ID", "Segment ID", "Segment Label", "Frequency", "Monetary Value", "Recency (Days)", "Avg Order Value", "Last Purchase"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.border
        row += 1
        
        # Customer rows
        for customer in customer_sample:
            ws.cell(row=row, column=1, value=customer.get("customer_id", ""))
            ws.cell(row=row, column=2, value=customer.get("segment_id", ""))
            ws.cell(row=row, column=3, value=customer.get("segment_label", ""))
            ws.cell(row=row, column=4, value=customer.get("frequency", 0))
            
            monetary = customer.get("monetary", 0)
            ws.cell(row=row, column=5, value=f"{monetary:,.2f}" if monetary else "0.00")
            
            ws.cell(row=row, column=6, value=customer.get("recency_days", 0))
            
            aov = customer.get("avg_order_value", 0)
            ws.cell(row=row, column=7, value=f"{aov:,.2f}" if aov else "0.00")
            
            last_purchase = customer.get("last_purchase_date", "")
            ws.cell(row=row, column=8, value=str(last_purchase) if last_purchase else "")
            
            for col_idx in range(1, 9):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1

    # =========================================================================
    # Future Agent Sheet Methods (Placeholder)
    # =========================================================================
    
    def _create_market_basket_sheet(self, wb: Workbook, agent_output: Dict[str, Any]):
        """Create market basket & sequence analysis sheet."""
        ws = wb.create_sheet("Market Basket & Sequence")
        self.styler.set_column_widths(ws, [28, 28, 18, 18, 18, 20, 45])
        
        row = 1
        ws[f'A{row}'] = "MARKET BASKET & SEQUENCE ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:G{row}')
        row += 2
        
        data = agent_output.get("data", {})
        analysis = data.get("analysis", {})
        parameters = analysis.get("parameters", {})
        columns_used = analysis.get("columns_used", {})
        mode = analysis.get("mode", "")
        
        # Metadata Section
        metadata = [
            ["Status", agent_output.get("status", "")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Mode", mode],
            ["Algorithm", analysis.get("algorithm", "")],
            ["Industry Context", analysis.get("industry", "")],
            ["Support Threshold", parameters.get("support", 0)],
            ["Confidence Threshold", parameters.get("confidence", 0)],
            ["Lift Threshold", parameters.get("lift", 0)],
            ["Gap Days", parameters.get("gap_days", 0)],
            ["Top N Rules", parameters.get("top_n_rules", 0)],
            ["Min Items / Transaction", parameters.get("min_items_per_transaction", 0)],
            ["Max Itemset Length", parameters.get("max_itemset_length", 0)],
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value) if value is not None else ""
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = self.subheader_fill
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Columns Used Section
        ws[f'A{row}'] = "COLUMNS USED"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        columns_info = [
            ["Transaction ID Column", columns_used.get("transaction_id_column", "")],
            ["Product ID Column", columns_used.get("product_id_column", "")],
            ["Customer ID Column", columns_used.get("customer_id_column", "")],
            ["Timestamp Column", columns_used.get("timestamp_column", "")],
        ]
        
        for key, value in columns_info:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value) if value is not None else ""
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 2
        
        # Mode-specific summary
        if mode == "within_basket":
            within = data.get("within_basket", {})
            summary_items = [
                ["Transactions", within.get("transactions", 0)],
                ["Unique Products", within.get("unique_products", 0)],
                ["Frequent Itemsets", len(within.get("frequent_itemsets", []))],
                ["Association Rules", len(within.get("association_rules", []))],
            ]
        else:
            cross = data.get("cross_transaction", {})
            summary_items = [
                ["Customers", cross.get("customers", 0)],
                ["Transactions", cross.get("transactions", 0)],
                ["Transitions", cross.get("total_transitions", 0)],
                ["Sequence Rules", len(cross.get("sequence_rules", []))],
            ]
        
        ws[f'A{row}'] = "SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        for key, value in summary_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 2
        
        if mode == "within_basket":
            # Frequent Itemsets
            ws[f'A{row}'] = "FREQUENT ITEMSETS"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            ws[f'A{row}'].fill = self.subheader_fill
            ws.merge_cells(f'A{row}:G{row}')
            row += 1
            
            headers = ["Items", "Size", "Count", "Support"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.center_alignment
                cell.border = self.border
            row += 1
            
            for itemset in data.get("within_basket", {}).get("frequent_itemsets", [])[:200]:
                ws.cell(row=row, column=1, value=", ".join(itemset.get("items", [])))
                ws.cell(row=row, column=2, value=itemset.get("size", 0))
                ws.cell(row=row, column=3, value=itemset.get("count", 0))
                ws.cell(row=row, column=4, value=itemset.get("support", 0))
                for col_idx in range(1, 5):
                    ws.cell(row=row, column=col_idx).border = self.border
                    ws.cell(row=row, column=col_idx).alignment = self.left_alignment
                row += 1
            
            row += 2
            
            # Association Rules
            ws[f'A{row}'] = "ASSOCIATION RULES"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            ws[f'A{row}'].fill = self.subheader_fill
            ws.merge_cells(f'A{row}:G{row}')
            row += 1
            
            headers = ["Antecedent", "Consequent", "Support", "Confidence", "Lift"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.center_alignment
                cell.border = self.border
            row += 1
            
            for rule in data.get("within_basket", {}).get("association_rules", [])[:200]:
                ws.cell(row=row, column=1, value=", ".join(rule.get("antecedent", [])))
                ws.cell(row=row, column=2, value=", ".join(rule.get("consequent", [])))
                ws.cell(row=row, column=3, value=rule.get("support", 0))
                ws.cell(row=row, column=4, value=rule.get("confidence", 0))
                ws.cell(row=row, column=5, value=rule.get("lift", 0))
                for col_idx in range(1, 6):
                    ws.cell(row=row, column=col_idx).border = self.border
                    ws.cell(row=row, column=col_idx).alignment = self.left_alignment
                row += 1
        else:
            # Sequence Rules
            ws[f'A{row}'] = "SEQUENCE RULES"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            ws[f'A{row}'].fill = self.subheader_fill
            ws.merge_cells(f'A{row}:G{row}')
            row += 1
            
            headers = ["From Item", "To Item", "Count", "Support", "Confidence", "Lift", "Avg Gap (Days)"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.center_alignment
                cell.border = self.border
            row += 1
            
            for rule in data.get("cross_transaction", {}).get("sequence_rules", [])[:200]:
                ws.cell(row=row, column=1, value=rule.get("from_item", ""))
                ws.cell(row=row, column=2, value=rule.get("to_item", ""))
                ws.cell(row=row, column=3, value=rule.get("count", 0))
                ws.cell(row=row, column=4, value=rule.get("support", 0))
                ws.cell(row=row, column=5, value=rule.get("confidence", 0))
                ws.cell(row=row, column=6, value=rule.get("lift", 0))
                ws.cell(row=row, column=7, value=rule.get("avg_gap_days", 0))
                for col_idx in range(1, 8):
                    ws.cell(row=row, column=col_idx).border = self.border
                    ws.cell(row=row, column=col_idx).alignment = self.left_alignment
                row += 1
        
    def _create_experiment_design_sheet(self, wb: Workbook, agent_output: Dict[str, Any]):
        """Create experimental design sheet."""
        ws = wb.create_sheet("Experimental Design")
        self.styler.set_column_widths(ws, [35, 25, 25, 50])
        
        row = 1
        ws[f'A{row}'] = "EXPERIMENTAL DESIGN ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        experiment_design = data.get("experiment_design", {})
        sample_calc = data.get("sample_size_calculation", {})
        feasibility = data.get("feasibility", {})
        quality = data.get("quality", {})
        
        # Metadata Section
        ws[f'A{row}'] = "ANALYSIS METADATA"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        metadata = [
            ["Status", agent_output.get("status", "")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Design Quality Score", quality.get("design_quality_score", 0)],
            ["Quality Status", quality.get("quality_status", "")],
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value) if value is not None else ""
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Experiment Design Parameters
        ws[f'A{row}'] = "EXPERIMENT DESIGN PARAMETERS"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        design_params = [
            ["Baseline Conversion Rate", f"{experiment_design.get('baseline_rate_pct', 0):.2f}%"],
            ["Minimum Detectable Lift", f"{experiment_design.get('min_detectable_lift_pct', 0):.1f}%"],
            ["Confidence Level", f"{experiment_design.get('confidence_pct', 0):.0f}%"],
            ["Statistical Power", f"{experiment_design.get('power_pct', 0):.0f}%"],
            ["Test Type", experiment_design.get("test_type", "").replace("_", " ").title()],
        ]
        
        for key, value in design_params:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value) if value is not None else ""
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Sample Size Calculation Results
        ws[f'A{row}'] = "SAMPLE SIZE CALCULATION"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        sample_results = [
            ["Control Group Size", f"{sample_calc.get('control_group_size', 0):,}"],
            ["Test Group Size", f"{sample_calc.get('test_group_size', 0):,}"],
            ["Total Sample Required", f"{sample_calc.get('total_sample_size', 0):,}"],
            ["Expected Control Rate", f"{sample_calc.get('expected_control_rate', 0)*100:.4f}%"],
            ["Expected Treatment Rate", f"{sample_calc.get('expected_treatment_rate', 0)*100:.4f}%"],
            ["Absolute Effect Size", f"{sample_calc.get('absolute_effect_size', 0):.6f}"],
            ["Z-Alpha (Two-Tailed)", f"{sample_calc.get('z_alpha', 0):.4f}"],
            ["Z-Beta", f"{sample_calc.get('z_beta', 0):.4f}"],
        ]
        
        for key, value in sample_results:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value) if value is not None else ""
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Feasibility Assessment
        ws[f'A{row}'] = "FEASIBILITY ASSESSMENT"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        feasibility_items = [
            ["Is Feasible", "Yes" if feasibility.get("is_feasible") else "No"],
            ["Feasibility Status", feasibility.get("status", "").upper()],
            ["Population Size", f"{feasibility.get('population_size', 0):,}" if feasibility.get('population_size') else "Not Provided"],
            ["Population Utilization", f"{feasibility.get('utilization_pct', 0):.1f}%" if feasibility.get('utilization_pct') else "N/A"],
            ["Feasibility Message", feasibility.get("message", "")],
        ]
        
        for key, value in feasibility_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value) if value is not None else ""
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            
            # Special handling for long message
            if key == "Feasibility Message":
                ws.merge_cells(f'B{row}:D{row}')
                ws[f'B{row}'].alignment = self.left_alignment
            row += 1
        
        row += 2
        
        # Visual Summary Table
        ws[f'A{row}'] = "EXPERIMENT SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        headers = ["Group", "Required Size", "Purpose"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.border
        row += 1
        
        summary_rows = [
            ["Control Group", f"{sample_calc.get('control_group_size', 0):,}", "Baseline experience (no change)"],
            ["Test Group", f"{sample_calc.get('test_group_size', 0):,}", "New experience (treatment)"],
            ["Total", f"{sample_calc.get('total_sample_size', 0):,}", "Combined experiment population"],
        ]
        
        for group_row in summary_rows:
            for col_idx, value in enumerate(group_row, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = self.border
                cell.alignment = self.left_alignment
            row += 1

    def _create_synthetic_control_sheet(self, wb: Workbook, agent_output: Dict[str, Any]):
        """Create synthetic control analysis sheet."""
        ws = wb.create_sheet("Synthetic Control")
        self.styler.set_column_widths(ws, [35, 25, 25, 25, 40])
        
        row = 1
        ws[f'A{row}'] = "SYNTHETIC CONTROL ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 2
        
        data = agent_output.get("data", {})
        summary_metrics = agent_output.get("summary_metrics", {})
        config = data.get("analysis_configuration", {})
        data_summary = data.get("data_summary", {})
        matching = data.get("matching_diagnostics", {})
        lift = data.get("lift_analysis", {})
        quality = data.get("quality", {})
        
        # Metadata Section
        ws[f'A{row}'] = "ANALYSIS METADATA"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        metadata = [
            ["Status", agent_output.get("status", "")],
            ["Execution Time (ms)", agent_output.get("execution_time_ms", 0)],
            ["Match Confidence Score", f"{quality.get('match_confidence_score', 0):.1f}%"],
            ["Quality Status", quality.get("quality_status", "")],
        ]
        
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value) if value is not None else ""
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Period Configuration
        periods = config.get("periods", {})
        ws[f'A{row}'] = "PERIOD CONFIGURATION"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        pre_period = periods.get("pre_period", {})
        treatment_period = periods.get("treatment_period", {})
        
        period_items = [
            ["Pre-Period Start", pre_period.get("start_date", "")],
            ["Pre-Period End", pre_period.get("end_date", "")],
            ["Pre-Period Days", pre_period.get("days", 0)],
            ["Treatment Start", treatment_period.get("start_date", "")],
            ["Treatment End", treatment_period.get("end_date", "")],
            ["Treatment Days", treatment_period.get("days", 0)],
        ]
        
        for key, value in period_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value) if value is not None else ""
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Data Summary
        ws[f'A{row}'] = "DATA SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        treatment_file = data_summary.get("treatment_file", {})
        baseline_file = data_summary.get("baseline_file", {})
        
        data_items = [
            ["Treatment File", treatment_file.get("filename", "")],
            ["Treatment Rows", treatment_file.get("valid_rows", 0)],
            ["Treatment Customers (Pre)", treatment_file.get("unique_customers_pre", 0)],
            ["Baseline File", baseline_file.get("filename", "")],
            ["Baseline Rows", baseline_file.get("valid_rows", 0)],
            ["Baseline Customers (Pre)", baseline_file.get("unique_customers_pre", 0)],
            ["Matched Control Customers", baseline_file.get("matched_customers", 0)],
        ]
        
        for key, value in data_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value) if value is not None else ""
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 1
        
        # Matching Diagnostics
        ws[f'A{row}'] = "MATCHING DIAGNOSTICS"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        match_items = [
            ["Treatment Count", matching.get("treatment_count", 0)],
            ["Control Pool Size", matching.get("control_pool_size", 0)],
            ["Matched Control Count", matching.get("matched_control_count", 0)],
            ["Avg Match Distance", f"{matching.get('avg_match_distance', 0):.4f}"],
            ["Max Match Distance", f"{matching.get('max_match_distance', 0):.4f}"],
            ["Match Confidence Score", f"{matching.get('match_confidence_score', 0):.1f}%"],
        ]
        
        for key, value in match_items:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value) if value is not None else ""
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1
        
        row += 2
        
        # Lift Analysis Results
        ws[f'A{row}'] = "LIFT ANALYSIS RESULTS"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        headers = ["Metric", "Treatment", "Synthetic Control", "Difference"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.border
        row += 1
        
        treatment_data = lift.get("treatment", {})
        control_data = lift.get("control", {})
        lift_data = lift.get("lift", {})
        
        lift_rows = [
            ["Pre-Period Avg", f"${treatment_data.get('pre_period_avg', 0):,.2f}", 
             f"${control_data.get('pre_period_avg', 0):,.2f}", 
             f"${treatment_data.get('pre_period_avg', 0) - control_data.get('pre_period_avg', 0):,.2f}"],
            ["Post-Period Avg", f"${treatment_data.get('post_period_avg', 0):,.2f}", 
             f"${control_data.get('post_period_avg', 0):,.2f}",
             f"${treatment_data.get('post_period_avg', 0) - control_data.get('post_period_avg', 0):,.2f}"],
            ["Change (Post - Pre)", f"${treatment_data.get('change_avg', 0):,.2f}", 
             f"${control_data.get('change_avg', 0):,.2f}",
             f"${lift_data.get('incremental_lift_per_customer', 0):,.2f}"],
            ["Customer Count", str(treatment_data.get('customer_count', 0)), 
             str(control_data.get('customer_count', 0)), "-"],
        ]
        
        for lift_row in lift_rows:
            for col_idx, value in enumerate(lift_row, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = self.border
                cell.alignment = self.left_alignment
            row += 1
        
        row += 1
        
        # Final Lift Summary
        ws[f'A{row}'] = "INCREMENTAL LIFT SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        lift_summary = [
            ["Incremental Lift / Customer", f"${lift_data.get('incremental_lift_per_customer', 0):,.2f}"],
            ["Total Incremental Lift", f"${lift_data.get('total_incremental_lift', 0):,.2f}"],
            ["Percentage Lift", f"{lift_data.get('percentage_lift', 0):.1f}%"],
            ["Counterfactual Post Avg", f"${lift_data.get('counterfactual_post_avg', 0):,.2f}"],
        ]
        
        for key, value in lift_summary:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value) if value is not None else ""
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = self.border
            ws[f'B{row}'].border = self.border
            row += 1

    def _create_synthetic_time_series_sheet(self, wb: Workbook, agent_output: Dict[str, Any]):
        """Create synthetic control time series data sheet."""
        ws = wb.create_sheet("Time Series Data")
        self.styler.set_column_widths(ws, [15, 18, 18, 15, 15])
        
        row = 1
        ws[f'A{row}'] = "TIME SERIES DATA"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 2
        
        data = agent_output.get("data", {})
        time_series = data.get("time_series", {})
        
        # Treatment Time Series
        ws[f'A{row}'] = "TREATMENT GROUP TIME SERIES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        headers = ["Date", "Daily Value", "Active Customers", "Transactions"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.border
        row += 1
        
        treatment_series = time_series.get("treatment", [])[:100]  # Limit for Excel
        for ts_row in treatment_series:
            ws.cell(row=row, column=1, value=ts_row.get("date", ""))
            ws.cell(row=row, column=2, value=f"${ts_row.get('value', 0):,.2f}")
            ws.cell(row=row, column=3, value=ts_row.get("active_customers", 0))
            ws.cell(row=row, column=4, value=ts_row.get("transactions", 0))
            for col_idx in range(1, 5):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
        
        row += 2
        
        # Control Time Series
        ws[f'A{row}'] = "SYNTHETIC CONTROL TIME SERIES"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.border
        row += 1
        
        control_series = time_series.get("synthetic_control", [])[:100]  # Limit for Excel
        for ts_row in control_series:
            ws.cell(row=row, column=1, value=ts_row.get("date", ""))
            ws.cell(row=row, column=2, value=f"${ts_row.get('value', 0):,.2f}")
            ws.cell(row=row, column=3, value=ts_row.get("active_customers", 0))
            ws.cell(row=row, column=4, value=ts_row.get("transactions", 0))
            for col_idx in range(1, 5):
                ws.cell(row=row, column=col_idx).border = self.border
                ws.cell(row=row, column=col_idx).alignment = self.left_alignment
            row += 1
